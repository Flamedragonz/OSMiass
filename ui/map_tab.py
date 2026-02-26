"""
Вкладка «Карта» — интерактивная карта на основе тайлов OsmAnd BigPlanet SQLite.

Слои рисуются снизу вверх; у каждого слоя — регулятор прозрачности и чекбокс.
Порядок слоёв (по умолчанию):
  1 — базовая карта (любой .sqlitedb, не распознанный как OSM / ФГИС)
  2 — OSM  (Miass_alpha.sqlitedb)
  3 — Кадастр ФГИС  (*fgis*.sqlitedb)

Формат BigPlanet: real_zoom = 17 − stored_z,  x/y — стандартные OSM-тайлы.
"""
import copy
import json
import math
import os
import re
import sqlite3
import uuid
import xml.etree.ElementTree as ET
from collections import deque
from dataclasses import dataclass, field
from typing import Optional, List, Tuple, Dict

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QSlider, QCheckBox, QGroupBox, QRadioButton, QButtonGroup,
    QApplication, QSizePolicy, QLineEdit, QTextEdit,
    QMenu, QColorDialog, QFileDialog, QComboBox, QToolButton,
    QTreeWidget, QTreeWidgetItem, QSplitter, QFrame,
)
from PyQt6.QtCore import Qt, pyqtSignal, QPointF, QRectF, QTimer
from PyQt6.QtGui import (
    QPixmap, QPainter, QColor, QFont, QFontMetrics, QPen, QImage,
    QBrush, QWheelEvent, QMouseEvent, QPolygonF, QPainterPath,
)

from core.config import Config

try:
    from shapely.geometry import Polygon as ShapelyPolygon
    from shapely.ops import unary_union
except ImportError:  # pragma: no cover - fallback для окружений без shapely
    ShapelyPolygon = None
    unary_union = None


# ══════════════════════════════════════════════════════════════
#  Тайловая математика (Slippy map / WGS-84)
# ══════════════════════════════════════════════════════════════

TILE_PX           = 256
BIGPLANET_OFFSET  = 17   # real_zoom = BIGPLANET_OFFSET − stored_z
MAX_OVERZOOM      = 4    # Уровней сверхзума сверх максимального тайлового зума


def stored_to_real(stored_z: int) -> int:
    return BIGPLANET_OFFSET - stored_z


def real_to_stored(real_z: int) -> int:
    return BIGPLANET_OFFSET - real_z


def lat_lon_to_tile_f(lat: float, lon: float, zoom: int) -> Tuple[float, float]:
    n     = 2.0 ** zoom
    x     = (lon + 180.0) / 360.0 * n
    lat_r = math.radians(lat)
    y     = (1.0 - math.log(math.tan(lat_r) + 1.0 / math.cos(lat_r)) / math.pi) / 2.0 * n
    return x, y


def tile_f_to_lat_lon(tx: float, ty: float, zoom: int) -> Tuple[float, float]:
    n     = 2.0 ** zoom
    lon   = tx / n * 360.0 - 180.0
    lat_r = math.atan(math.sinh(math.pi * (1.0 - 2.0 * ty / n)))
    return math.degrees(lat_r), lon


def haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Расстояние между двумя точками (метры), формула Хаверсина."""
    R    = 6_371_000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dl   = math.radians(lon2 - lon1)
    a    = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dl / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ══════════════════════════════════════════════════════════════
#  Система точек
# ══════════════════════════════════════════════════════════════

PRESET_ICONS    = ("circle", "square", "diamond", "triangle", "star", "pin", "cross")
PRESET_ICONS_RU = ("Круг",   "Квадрат", "Ромб",   "Треугольник", "Звезда", "Маркер", "Крест")
BASE_POINT_PX   = 16   # базовый диаметр иконки при size = 1.0


@dataclass
class MapPoint:
    lat:              float
    lon:              float
    label:            str   = ""
    icon_type:        str   = "circle"    # из PRESET_ICONS или "custom"
    color:            str   = "#89b4fa"
    size:             float = 1.0         # множитель 0.5 – 3.0
    custom_icon_path: str   = ""
    id:               str   = field(default_factory=lambda: uuid.uuid4().hex[:8])
    is_intersection:  bool  = False       # True = подтверждённая точка пересечения
    src_edge_a_endpoints: tuple = field(default_factory=tuple)  # (id1, id2) исходного ребра A
    src_edge_b_endpoints: tuple = field(default_factory=tuple)  # (id1, id2) исходного ребра B


EDGE_TYPES    = ("solid", "dashed")
EDGE_TYPES_RU = ("Сплошная", "Пунктир")


@dataclass
class MapEdge:
    point_a_id: str
    point_b_id: str
    color:     str   = "#a6adc8"
    width:     float = 2.0
    line_type: str   = "solid"    # "solid" | "dashed"
    opacity:   float = 1.0
    id:        str   = field(default_factory=lambda: uuid.uuid4().hex[:8])


@dataclass
class MapPolygon:
    point_ids:    List[str]         # упорядоченные вершины
    fill_color:   str   = "#89b4fa"
    fill_opacity: float = 0.25
    border_color: str   = "#89b4fa"
    border_width: float = 1.5
    border_type:  str   = "solid"   # "solid" | "dashed"
    id:           str   = field(default_factory=lambda: uuid.uuid4().hex[:8])


@dataclass
class DynamicPoint:
    """Временная точка пересечения двух рёбер (не хранится в map_points)."""
    edge_a_id: str
    edge_b_id: str
    id: str = field(default_factory=lambda: uuid.uuid4().hex[:8])


def _make_star_points(
    cx: float, cy: float, r_outer: float, r_inner: float, n: int = 5
) -> QPolygonF:
    """Возвращает QPolygonF n-конечной звезды."""
    pts = []
    for i in range(2 * n):
        angle = math.radians(-90.0 + i * 180.0 / n)
        r     = r_outer if i % 2 == 0 else r_inner
        pts.append(QPointF(cx + r * math.cos(angle), cy + r * math.sin(angle)))
    return QPolygonF(pts)


# ══════════════════════════════════════════════════════════════
#  Дескриптор слоя
# ══════════════════════════════════════════════════════════════

class LayerInfo:
    """Один слой карты: провайдер тайлов + параметры отображения."""

    __slots__ = ("provider", "name", "opacity", "visible")

    def __init__(self, provider, name: str,
                 opacity: float = 1.0, visible: bool = True):
        self.provider = provider
        self.name     = name
        self.opacity  = opacity  # 0.0 – 1.0
        self.visible  = visible


# ══════════════════════════════════════════════════════════════
#  Провайдер тайлов (SQLite BigPlanet)
# ══════════════════════════════════════════════════════════════

class TileProvider:
    """Загружает тайлы из SQLite-базы BigPlanet OsmAnd."""

    MAX_CACHE = 800

    def __init__(self, db_path: str, name: str = ""):
        self.db_path = db_path
        self.name    = name or os.path.basename(db_path)
        self._cache: Dict[tuple, Optional[QPixmap]] = {}
        self._conn   = sqlite3.connect(db_path, check_same_thread=False)
        self._real_zoom_range = self._detect_zoom_range()

    def close(self):
        try:
            self._conn.close()
        except Exception:
            pass

    # ── Диапазон зумов ──────────────────────────────────────────

    def _detect_zoom_range(self) -> Tuple[int, int]:
        try:
            cur = self._conn.cursor()
            cur.execute("SELECT MIN(z), MAX(z) FROM tiles")
            mn, mx = cur.fetchone()
            if mn is None:
                return (7, 15)
            return (stored_to_real(mx), stored_to_real(mn))
        except Exception:
            return (7, 15)

    @property
    def min_zoom(self) -> int:
        return self._real_zoom_range[0]

    @property
    def max_zoom(self) -> int:
        return self._real_zoom_range[1]

    # ── Центр покрытия ──────────────────────────────────────────

    def coverage_center(self, hint_zoom: int = 12) -> Optional[Tuple[float, float]]:
        stored_z = real_to_stored(hint_zoom)
        cur = self._conn.cursor()
        for offset in (0, 1, -1, 2, -2, 3, -3):
            sz = stored_z + offset
            if sz < 0:
                continue
            cur.execute(
                "SELECT AVG(x), AVG(y), COUNT(*) FROM tiles WHERE z=?", (sz,)
            )
            row = cur.fetchone()
            if row and row[2] and row[2] > 0:
                rz  = stored_to_real(sz)
                lat, lon = tile_f_to_lat_lon(row[0] + 0.5, row[1] + 0.5, rz)
                return lat, lon
        return None

    # ── Загрузка тайла из БД ─────────────────────────────────────

    def _load_from_db(
        self, tile_x: int, tile_y: int, real_zoom: int
    ) -> Optional[QPixmap]:
        """Непосредственная загрузка тайла из SQLite без кэширования."""
        stored_z = real_to_stored(real_zoom)
        try:
            cur = self._conn.cursor()
            cur.execute(
                "SELECT image FROM tiles WHERE x=? AND y=? AND z=?",
                (tile_x, tile_y, stored_z),
            )
            row = cur.fetchone()
            if row and row[0]:
                data = bytes(row[0])
                qimg = QImage.fromData(data)
                if not qimg.isNull():
                    px = QPixmap.fromImage(qimg)
                    if not px.isNull():
                        return px
        except Exception:
            pass
        return None

    # ── Загрузка тайла с поддержкой сверхзума ───────────────────

    def get_pixmap(
        self, tile_x: int, tile_y: int, real_zoom: int
    ) -> Optional[QPixmap]:
        key = (tile_x, tile_y, real_zoom)
        if key in self._cache:
            return self._cache[key]

        if real_zoom <= self.max_zoom:
            # Обычная загрузка тайла
            px = self._load_from_db(tile_x, tile_y, real_zoom)
        else:
            # Сверхзум: берём родительский тайл и масштабируем фрагмент
            zoom_diff  = min(real_zoom - self.max_zoom, MAX_OVERZOOM)
            fetch_zoom = real_zoom - zoom_diff    # == self.max_zoom
            scale      = 1 << zoom_diff           # 2^zoom_diff

            parent_x  = tile_x >> zoom_diff
            parent_y  = tile_y >> zoom_diff

            # Сначала проверяем кэш родительского тайла
            parent_key = (parent_x, parent_y, fetch_zoom)
            parent_px  = self._cache.get(parent_key)
            if parent_px is None:
                parent_px = self._load_from_db(parent_x, parent_y, fetch_zoom)
                # Кэшируем родительский тайл отдельно
                if len(self._cache) >= self.MAX_CACHE:
                    try:
                        del self._cache[next(iter(self._cache))]
                    except StopIteration:
                        pass
                self._cache[parent_key] = parent_px

            if parent_px:
                sub_x     = tile_x & (scale - 1)
                sub_y     = tile_y & (scale - 1)
                crop_size = TILE_PX // scale
                crop_x    = sub_x * crop_size
                crop_y    = sub_y * crop_size
                cropped   = parent_px.copy(crop_x, crop_y, crop_size, crop_size)
                px = cropped.scaled(
                    TILE_PX, TILE_PX,
                    Qt.AspectRatioMode.IgnoreAspectRatio,
                    Qt.TransformationMode.FastTransformation,
                )
            else:
                px = None

        # LRU-вытеснение
        if len(self._cache) >= self.MAX_CACHE:
            try:
                del self._cache[next(iter(self._cache))]
            except StopIteration:
                pass

        self._cache[key] = px
        return px

    def has_tile(self, tile_x: int, tile_y: int, real_zoom: int) -> bool:
        stored_z = real_to_stored(real_zoom)
        try:
            cur = self._conn.cursor()
            cur.execute(
                "SELECT 1 FROM tiles WHERE x=? AND y=? AND z=? LIMIT 1",
                (tile_x, tile_y, stored_z),
            )
            return cur.fetchone() is not None
        except Exception:
            return False


# ══════════════════════════════════════════════════════════════
#  Холст карты
# ══════════════════════════════════════════════════════════════

class MapCanvas(QWidget):
    """Интерактивный виджет карты с поддержкой зума и панорамирования."""

    coord_clicked    = pyqtSignal(float, float)  # lat, lon при клике (только в режиме редактирования)
    coord_hovered    = pyqtSignal(float, float)  # lat, lon при движении мыши
    temp_pts_changed = pyqtSignal()              # при изменении временных точек

    # Сигналы системы точек
    point_added       = pyqtSignal(object)   # MapPoint
    point_moved       = pyqtSignal(object)   # MapPoint
    point_deleted     = pyqtSignal(object)   # MapPoint
    selection_changed = pyqtSignal(object)   # List[MapPoint]
    edge_added        = pyqtSignal(object)   # MapEdge
    edge_deleted      = pyqtSignal(object)   # MapEdge
    labels_toggled    = pyqtSignal(bool)     # при переключении глобальных меток (H)

    def __init__(self, parent=None):
        super().__init__(parent)

        # Слои карты (снизу вверх)
        self.layers: List[LayerInfo] = []

        # Состояние карты
        self.center_lat = 55.05
        self.center_lon = 60.10
        self.zoom       = 12

        # Маркеры данных: (lat, lon, label, color_hex)
        self.markers: List[Tuple[float, float, str, str]] = []
        self.selected_marker_idx: Optional[int] = None

        # Режим обзора/измерения
        self.view_mode   = False                           # True = обзор, False = редактирование
        self.temp_points: List[Tuple[float, float]] = []  # Временные точки измерения

        # Система точек, рёбер и полигонов
        self.map_points:   List[MapPoint]   = []
        self.map_edges:    List[MapEdge]    = []
        self.map_polygons: List[MapPolygon] = []
        self.point_mode:   bool             = False
        self.auto_polygon: bool             = True   # авто-замыкание полигонов
        self.show_geo_labels: bool          = True   # метки расстояний/углов на рёбрах
        self.show_coord_labels: bool        = True
        self.show_angle_labels: bool        = True
        self.show_area_labels: bool         = True
        self.show_edge_length_labels: bool  = True
        self.area_unit:    str              = "m2"   # "m2" | "ha"

        # Выделение (multi-select)
        self.selected_point_ids: set          = set()       # set[str]
        self._hovered_point_id:  Optional[str] = None
        self._pivot_point_id:    Optional[str] = None
        self._selected_polygon_id: Optional[str] = None

        # Иконки (кэш)
        self._icon_cache: Dict[str, Optional[QPixmap]] = {}

        # Перетаскивание группы точек
        self._dragging_point:    Optional[MapPoint]              = None
        self._drag_was_moved:    bool                             = False
        self._drag_start_screen: Optional[QPointF]               = None
        self._group_drag_origin: Dict[str, Tuple[float, float]]  = {}
        self._dragging_polygon: Optional[MapPolygon] = None

        # Рамка выделения (rubber band)
        self._rubber_band_start:  Optional[QPointF] = None
        self._rubber_band_end:    Optional[QPointF] = None
        self._rubber_band_active: bool               = False

        # Режим ручного соединения рёбрами
        self._connect_mode:      bool          = False
        self._connect_source_id: Optional[str] = None

        # Настройки по умолчанию для новых точек
        self.default_icon_type = "circle"
        self.default_color     = "#89b4fa"
        self.default_size      = 1.0

        # Настройки по умолчанию для новых рёбер
        self.default_edge_color   = "#a6adc8"
        self.default_edge_width   = 2.0
        self.default_edge_type    = "solid"
        self.default_edge_opacity = 1.0
        self.auto_connect         = False

        # Настройки по умолчанию для новых полигонов
        self.default_poly_fill         = "#89b4fa"
        self.default_poly_fill_opacity = 0.25
        self.default_poly_border       = "#89b4fa"
        self.default_poly_border_width = 1.5
        self.default_poly_border_type  = "solid"

        # Визуализация / глобальные метки (H-key)
        self._hovered_edge_id: Optional[str] = None
        self.show_point_labels: bool = False   # True = показывать тултип для всех точек

        # Undo / Redo
        self._undo_stack: List[dict] = []
        self._redo_stack: List[dict] = []
        _UNDO_MAX = 50

        # Авто-соединение: последняя и первая точка текущей цепи
        self._last_placed_point_id: Optional[str] = None
        self._chain_start_id:       Optional[str] = None

        # Динамические точки пересечений рёбер
        self.dynamic_points: List[DynamicPoint] = []
        self.show_dynamic_points: bool = True
        self.auto_dynamic_points: bool = True

        # R-поворот мышью
        self._rotate_mode:          bool                       = False
        self._rotate_screen_ref:    Optional[Tuple[float,float]] = None
        self._rotate_center_screen: Optional[Tuple[float,float]] = None
        self._rotate_center_latlon: Optional[Tuple[float,float]] = None
        self._rotate_origins:       Dict[str, Tuple[float,float]] = {}

        # Панорамирование
        self._drag_pos:    Optional[QPointF]           = None
        self._drag_center: Optional[Tuple[float, float]] = None

        self.setMouseTracking(True)
        self.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
        self.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding
        )
        self.setMinimumSize(400, 300)

        self._update_timer = QTimer(self)
        self._update_timer.setSingleShot(True)
        self._update_timer.timeout.connect(self.update)

    # ── Координатные вычисления ──────────────────────────────────

    def _center_tile(self) -> Tuple[float, float]:
        return lat_lon_to_tile_f(self.center_lat, self.center_lon, self.zoom)

    def _screen_to_lat_lon(self, sx: float, sy: float) -> Tuple[float, float]:
        cx, cy = self._center_tile()
        W, H   = self.width(), self.height()
        tx = cx + (sx - W / 2.0) / TILE_PX
        ty = cy + (sy - H / 2.0) / TILE_PX
        return tile_f_to_lat_lon(tx, ty, self.zoom)

    def _lat_lon_to_screen(self, lat: float, lon: float) -> Tuple[int, int]:
        tx, ty = lat_lon_to_tile_f(lat, lon, self.zoom)
        cx, cy = self._center_tile()
        W, H   = self.width(), self.height()
        sx = int((tx - cx) * TILE_PX + W / 2.0)
        sy = int((ty - cy) * TILE_PX + H / 2.0)
        return sx, sy

    # ── Публичные методы ─────────────────────────────────────────

    def set_center(self, lat: float, lon: float, zoom: int | None = None):
        self.center_lat = lat
        self.center_lon = lon
        if zoom is not None:
            self.zoom = zoom
        self.update()

    def set_zoom(self, zoom: int):
        if self.layers:
            all_p = [l.provider for l in self.layers]
            min_z = min(p.min_zoom for p in all_p)
            max_z = max(p.max_zoom for p in all_p)
            zoom  = max(min_z, min(max_z + MAX_OVERZOOM, zoom))
        self.zoom = zoom
        self.update()

    def update_markers(self, rows: list):
        self.markers.clear()
        for i, row in enumerate(rows):
            lat_s = row.get("широта",  "").strip()
            lon_s = row.get("долгота", "").strip()
            if lat_s and lon_s:
                try:
                    lat   = float(lat_s.replace(",", "."))
                    lon   = float(lon_s.replace(",", "."))
                    label = row.get("выдел", f"{i+1}")
                    self.markers.append((lat, lon, label, "#f38ba8"))
                except ValueError:
                    pass
        self.update()

    # ── Отрисовка ────────────────────────────────────────────────

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        W, H = self.width(), self.height()

        # Фон
        painter.fillRect(0, 0, W, H, QColor("#27273a"))

        if not self.layers:
            painter.setPen(QColor("#a6adc8"))
            painter.setFont(QFont("Segoe UI", 12))
            painter.drawText(
                QRectF(0, 0, W, H),
                Qt.AlignmentFlag.AlignCenter,
                "Карта не загружена\n\nПоместите файлы .sqlitedb в папку osm_map/",
            )
            return

        cx, cy = self._center_tile()
        half_x = int(W / (2 * TILE_PX)) + 2
        half_y = int(H / (2 * TILE_PX)) + 2
        t_x0   = int(cx) - half_x
        t_y0   = int(cy) - half_y
        t_x1   = int(cx) + half_x + 1
        t_y1   = int(cy) + half_y + 1
        n      = 2 ** self.zoom

        visible_layers = [l for l in self.layers if l.visible]

        for tx in range(t_x0, t_x1):
            for ty in range(t_y0, t_y1):
                if tx < 0 or ty < 0 or tx >= n or ty >= n:
                    continue

                sx = int((tx - cx) * TILE_PX + W / 2.0)
                sy = int((ty - cy) * TILE_PX + H / 2.0)

                if sx > W or sy > H or sx + TILE_PX < 0 or sy + TILE_PX < 0:
                    continue

                any_drawn = False
                for layer in visible_layers:
                    # Разрешаем рендер при сверхзуме (max_zoom проверка убрана —
                    # get_pixmap сам вернёт масштабированный родительский тайл)
                    if layer.provider.min_zoom > self.zoom:
                        continue
                    px = layer.provider.get_pixmap(tx, ty, self.zoom)
                    if px and not px.isNull():
                        painter.setOpacity(layer.opacity)
                        painter.drawPixmap(sx, sy, TILE_PX, TILE_PX, px)
                        any_drawn = True

                painter.setOpacity(1.0)
                if not any_drawn:
                    painter.fillRect(sx, sy, TILE_PX, TILE_PX, QColor("#1e1e2e"))
                    painter.setPen(QPen(QColor("#313244"), 1))
                    painter.drawRect(sx, sy, TILE_PX - 1, TILE_PX - 1)

        self._draw_map_polygons(painter)
        self._draw_map_edges(painter)
        self._draw_edge_labels(painter)
        self._draw_markers(painter)
        self._draw_map_points(painter)
        self._draw_dynamic_points(painter)
        self._draw_polygon_labels(painter)
        self._draw_rubber_band(painter)
        self._draw_point_labels(painter)
        self._draw_temp_measurements(painter)
        self._draw_crosshair(painter, W, H)
        self._draw_scale(painter, W, H)

        # ─── Индикатор текущего режима (верхний правый угол) ──────
        font_mode = QFont("Segoe UI", 9, QFont.Weight.Bold)
        fm_mode   = QFontMetrics(font_mode)

        def _draw_badge(text: str, color: str, y_off: int = 8):
            tw  = fm_mode.horizontalAdvance(text)
            th  = fm_mode.height()
            bx  = W - tw - 16
            painter.setOpacity(0.72)
            painter.fillRect(bx - 6, y_off, tw + 12, th + 6, QColor("#000000"))
            painter.setOpacity(1.0)
            painter.setFont(font_mode)
            painter.setPen(QColor(color))
            painter.drawText(bx, y_off + th, text)
            return th + 6 + 4   # высота + отступ

        y = 8
        if self.view_mode:
            y += _draw_badge("Обзор / Измерение", "#f9e2af", y)
        if self.point_mode:
            if self._rotate_mode:
                sub = "↻ Поворот [R]"
                col = "#f38ba8"
            elif self._connect_mode:
                sub = "● Соединение"
                col = "#a6e3a1"
            else:
                sub = "✎ Точки  [Ctrl+ЛКМ=добавить]"
                col = "#89b4fa"
            y += _draw_badge(sub, col, y)
            h_hint = "[H] Метки вкл" if self.show_geo_labels else "[H] Метки выкл"
            _draw_badge(h_hint, "#6c7086", y)

        painter.end()

    def _draw_markers(self, painter: QPainter):
        W, H = self.width(), self.height()
        for idx, (lat, lon, label, color) in enumerate(self.markers):
            sx, sy = self._lat_lon_to_screen(lat, lon)
            if sx < -20 or sy < -20 or sx > W + 20 or sy > H + 20:
                continue

            is_sel = (idx == self.selected_marker_idx)
            c = QColor(color)
            r = 9 if is_sel else 7

            painter.setOpacity(1.0)
            painter.setPen(QPen(QColor("#1e1e2e"), 2))
            painter.setBrush(QBrush(c))
            painter.drawEllipse(sx - r, sy - r, r * 2, r * 2)

            if is_sel:
                painter.setPen(QPen(QColor("#cba6f7"), 2))
                painter.setBrush(Qt.BrushStyle.NoBrush)
                painter.drawEllipse(sx - r - 3, sy - r - 3, (r + 3) * 2, (r + 3) * 2)

            painter.setPen(QPen(QColor("#cdd6f4"), 1))
            painter.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
            painter.drawText(sx + r + 3, sy + 4, label)

    def _draw_map_polygons(self, painter: QPainter):
        """Рисует закрашенные полигоны под рёбрами."""
        for poly in self.map_polygons:
            pts = [self._get_point_by_id(pid) for pid in poly.point_ids]
            if any(p is None for p in pts) or len(pts) < 3:
                continue
            screen = [QPointF(*self._lat_lon_to_screen(p.lat, p.lon)) for p in pts]
            path = QPainterPath()
            path.moveTo(screen[0])
            for sp in screen[1:]:
                path.lineTo(sp)
            path.closeSubpath()

            is_sel_poly = (poly.id == self._selected_polygon_id)
            painter.setOpacity(poly.fill_opacity + (0.1 if is_sel_poly else 0.0))
            painter.fillPath(path, QBrush(QColor(poly.fill_color)))

            pen = QPen(QColor(poly.border_color), poly.border_width + (1.0 if is_sel_poly else 0.0))
            pen.setCapStyle(Qt.PenCapStyle.RoundCap)
            if poly.border_type == "dashed":
                pen.setStyle(Qt.PenStyle.CustomDashLine)
                pen.setDashPattern([6.0, 4.0])
            painter.setOpacity(1.0)
            painter.setPen(pen)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawPath(path)

    def _draw_edge_labels(self, painter: QPainter):
        """Рисует метки расстояний и дирекционных углов на рёбрах.
        При show_geo_labels=True — все рёбра. Иначе — только hovered."""
        if not self.map_edges:
            return
        has_hover = self._hovered_edge_id is not None
        if not self.show_geo_labels and not has_hover:
            return

        font     = QFont("Segoe UI", 7)
        font_hov = QFont("Segoe UI", 8, QFont.Weight.Bold)
        fm_n     = QFontMetrics(font)
        fm_h     = QFontMetrics(font_hov)
        W, H     = self.width(), self.height()

        for edge in self.map_edges:
            is_hov = (edge.id == self._hovered_edge_id)
            if not self.show_geo_labels and not is_hov:
                continue
            pt_a = self._get_point_by_id(edge.point_a_id)
            pt_b = self._get_point_by_id(edge.point_b_id)
            if pt_a is None or pt_b is None:
                continue
            sx1, sy1 = self._lat_lon_to_screen(pt_a.lat, pt_a.lon)
            sx2, sy2 = self._lat_lon_to_screen(pt_b.lat, pt_b.lon)
            mx_e, my_e = (sx1 + sx2) // 2, (sy1 + sy2) // 2
            if mx_e < -60 or mx_e > W + 60 or my_e < -60 or my_e > H + 60:
                continue

            parts = []
            if self.show_edge_length_labels:
                dist = haversine_distance(pt_a.lat, pt_a.lon, pt_b.lat, pt_b.lon)
                dist_str = (f"{dist/1000:.3f} км" if dist >= 1000 else
                            f"{dist:.2f} м"       if dist >= 1   else
                            f"{dist*100:.1f} см")
                parts.append(dist_str)
            if self.show_angle_labels:
                bearing = self._calc_bearing(pt_a.lat, pt_a.lon, pt_b.lat, pt_b.lon)
                parts.append(f"∠{bearing:.1f}°")
            if not parts:
                continue
            lbl = "  ".join(parts)

            f   = font_hov if is_hov else font
            fm  = fm_h     if is_hov else fm_n
            tw  = fm.horizontalAdvance(lbl) + 8
            th  = fm.height()
            bx  = mx_e - tw // 2
            by  = my_e - th // 2 - (4 if is_hov else 0)

            bg_color  = QColor("#313244") if is_hov else QColor("#1e1e2e")
            txt_color = QColor("#f9e2af") if is_hov else QColor("#cdd6f4")

            painter.setOpacity(0.85 if is_hov else 0.72)
            painter.fillRect(bx, by, tw, th + 2, bg_color)
            if is_hov:
                painter.setOpacity(0.9)
                painter.setPen(QPen(QColor("#89b4fa"), 1))
                painter.drawRect(bx, by, tw, th + 2)
            painter.setOpacity(1.0)
            painter.setFont(f)
            painter.setPen(txt_color)
            painter.drawText(bx + 4, by + th - 1, lbl)

    def _draw_point_labels(self, painter: QPainter):
        """Рисует тултипы точек: имя / координаты / дирекционный угол.
        Всегда для hovered-точки; для всех — при show_point_labels=True."""
        if not self.point_mode or not self.map_points:
            return
        has_hover = self._hovered_point_id is not None
        if not self.show_point_labels and not has_hover:
            return

        font_hdr = QFont("Segoe UI", 9, QFont.Weight.Bold)
        font_val = QFont("Consolas", 8)
        fm_h     = QFontMetrics(font_hdr)
        fm_v     = QFontMetrics(font_val)
        line_h   = max(fm_h.height(), fm_v.height()) + 2
        pad      = 5
        W, H     = self.width(), self.height()

        for pt in self.map_points:
            is_hov = (pt.id == self._hovered_point_id)
            if not self.show_point_labels and not is_hov:
                continue
            sx, sy = self._lat_lon_to_screen(pt.lat, pt.lon)
            if sx < -80 or sx > W + 80 or sy < -80 or sy > H + 80:
                continue

            # Строки тултипа
            lines: List[Tuple[str, bool]] = []   # (text, is_header)
            if pt.label:
                lines.append((pt.label, True))
            if self.show_coord_labels:
                lines.append((f"{pt.lat:+.6f}° N", False))
                lines.append((f"{pt.lon:+.6f}° E", False))

            # Дирекционный угол к первому смежному ребру
            for e in self.map_edges:
                other_id = None
                if e.point_a_id == pt.id:
                    other_id = e.point_b_id
                elif e.point_b_id == pt.id:
                    other_id = e.point_a_id
                if other_id:
                    other = self._get_point_by_id(other_id)
                    if other and self.show_angle_labels:
                        bear = self._calc_bearing(pt.lat, pt.lon, other.lat, other.lon)
                        lines.append((f"∠ {bear:.1f}°", False))
                    break

            if not lines:
                continue

            # Ширина блока по максимальной строке
            box_w = max(
                (fm_h if hdr else fm_v).horizontalAdvance(txt)
                for txt, hdr in lines
            ) + pad * 2
            box_h = len(lines) * line_h + pad * 2

            # Автопозиционирование: предпочитаем верхний-правый квадрант
            r_off = max(4, int(BASE_POINT_PX * pt.size / 2)) + 8
            ox = sx + r_off
            oy = sy - box_h - 4
            if ox + box_w > W:
                ox = sx - box_w - r_off
            if oy < 0:
                oy = sy + r_off
            if oy + box_h > H:
                oy = H - box_h - 4
            ox = max(0, ox)

            # Фон
            painter.setOpacity(0.88 if is_hov else 0.72)
            painter.fillRect(int(ox), int(oy), box_w, box_h, QColor("#1e1e2e"))
            painter.setOpacity(0.9 if is_hov else 0.6)
            painter.setPen(QPen(QColor("#89b4fa" if is_hov else "#45475a"), 1))
            painter.drawRect(int(ox), int(oy), box_w, box_h)

            # Текст
            painter.setOpacity(1.0)
            for i, (txt, hdr) in enumerate(lines):
                ty = int(oy) + pad + (i + 1) * line_h - 2
                painter.setFont(font_hdr if hdr else font_val)
                painter.setPen(QColor("#f9e2af" if hdr else "#cdd6f4"))
                painter.drawText(int(ox) + pad, ty, txt)

    def _draw_polygon_labels(self, painter: QPainter):
        """Рисует метки площади в центроиде каждого полигона."""
        if not self.show_geo_labels or not self.show_area_labels or not self.map_polygons:
            return
        font = QFont("Segoe UI", 8, QFont.Weight.Bold)
        fm   = QFontMetrics(font)
        W, H = self.width(), self.height()
        for poly in self.map_polygons:
            pts = [self._get_point_by_id(pid) for pid in poly.point_ids]
            if any(p is None for p in pts) or len(pts) < 3:
                continue
            area  = self._effective_polygon_area(poly)
            if self.area_unit == "ha":
                lbl = f"{area/10000:.4f} га"
            else:
                lbl = f"{area:.1f} м²"
            screen = [self._lat_lon_to_screen(p.lat, p.lon) for p in pts]
            cx = sum(s[0] for s in screen) // len(screen)
            cy = sum(s[1] for s in screen) // len(screen)
            if cx < -60 or cx > W + 60 or cy < -60 or cy > H + 60:
                continue
            tw = fm.horizontalAdvance(lbl) + 10
            th = fm.height()
            bx = cx - tw // 2
            by = cy - th // 2
            painter.setOpacity(0.78)
            painter.fillRect(bx, by, tw, th + 4, QColor("#1e1e2e"))
            painter.setOpacity(1.0)
            painter.setFont(font)
            painter.setPen(QColor("#a6e3a1"))
            painter.drawText(bx + 5, by + th, lbl)

    def _draw_map_edges(self, painter: QPainter):
        """Рисует рёбра (линии) между точками."""
        W, H = self.width(), self.height()
        for edge in self.map_edges:
            pt_a = self._get_point_by_id(edge.point_a_id)
            pt_b = self._get_point_by_id(edge.point_b_id)
            if pt_a is None or pt_b is None:
                continue
            sx1, sy1 = self._lat_lon_to_screen(pt_a.lat, pt_a.lon)
            sx2, sy2 = self._lat_lon_to_screen(pt_b.lat, pt_b.lon)
            # Грубая отсечка
            if max(sx1, sx2) < -50 or min(sx1, sx2) > W + 50:
                continue
            if max(sy1, sy2) < -50 or min(sy1, sy2) > H + 50:
                continue
            pen = QPen(QColor(edge.color), edge.width)
            pen.setCapStyle(Qt.PenCapStyle.RoundCap)
            if edge.line_type == "dashed":
                pen.setStyle(Qt.PenStyle.CustomDashLine)
                pen.setDashPattern([6.0, 4.0])
            else:
                pen.setStyle(Qt.PenStyle.SolidLine)
            painter.setOpacity(edge.opacity)
            painter.setPen(pen)
            painter.drawLine(sx1, sy1, sx2, sy2)
        painter.setOpacity(1.0)

    def _draw_rubber_band(self, painter: QPainter):
        """Рисует рамку множественного выделения."""
        if not self._rubber_band_active:
            return
        if self._rubber_band_start is None or self._rubber_band_end is None:
            return
        x1 = int(min(self._rubber_band_start.x(), self._rubber_band_end.x()))
        y1 = int(min(self._rubber_band_start.y(), self._rubber_band_end.y()))
        x2 = int(max(self._rubber_band_start.x(), self._rubber_band_end.x()))
        y2 = int(max(self._rubber_band_start.y(), self._rubber_band_end.y()))
        w, h = x2 - x1, y2 - y1
        if w < 2 or h < 2:
            return
        painter.setOpacity(0.15)
        painter.fillRect(x1, y1, w, h, QColor("#89b4fa"))
        painter.setOpacity(1.0)
        painter.setPen(QPen(QColor("#89b4fa"), 1, Qt.PenStyle.DashLine))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRect(x1, y1, w, h)

    def _draw_map_points(self, painter: QPainter):
        """Рисует пользовательские точки (система точек)."""
        W, H = self.width(), self.height()
        for pt in self.map_points:
            sx, sy = self._lat_lon_to_screen(pt.lat, pt.lon)
            if sx < -60 or sy < -60 or sx > W + 60 or sy > H + 60:
                continue
            is_sel = (pt.id in self.selected_point_ids)
            is_hov = (pt.id == self._hovered_point_id)
            self._draw_point_icon(painter, sx, sy, pt, is_sel, is_hov)
            # Опорная точка для поворота
            if pt.id == self._pivot_point_id:
                r2 = max(4, int(BASE_POINT_PX * pt.size / 2)) + 6
                painter.setOpacity(0.9)
                painter.setPen(QPen(QColor("#f38ba8"), 1.5))
                painter.setBrush(Qt.BrushStyle.NoBrush)
                painter.drawEllipse(sx - r2, sy - r2, 2 * r2, 2 * r2)
                painter.drawLine(sx - 4, sy, sx + 4, sy)
                painter.drawLine(sx, sy - 4, sx, sy + 4)
            # Индикатор источника в режиме ручного соединения
            if self._connect_mode and pt.id == self._connect_source_id:
                r2 = max(4, int(BASE_POINT_PX * pt.size / 2)) + 8
                painter.setOpacity(0.8)
                painter.setPen(QPen(QColor("#a6e3a1"), 1.5, Qt.PenStyle.DashLine))
                painter.setBrush(Qt.BrushStyle.NoBrush)
                painter.drawEllipse(sx - r2, sy - r2, 2 * r2, 2 * r2)
            if pt.label:
                r = max(4, int(BASE_POINT_PX * pt.size / 2))
                painter.setOpacity(1.0)
                painter.setPen(QPen(QColor("#cdd6f4"), 1))
                painter.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
                painter.drawText(sx + r + 4, sy + 4, pt.label)

    def _draw_point_icon(
        self, painter: QPainter,
        cx: int, cy: int, pt: MapPoint,
        is_sel: bool, is_hov: bool,
    ):
        r     = max(4, int(BASE_POINT_PX * pt.size / 2))
        color = QColor(pt.color)
        painter.setOpacity(1.0)

        if pt.icon_type == "custom" and pt.custom_icon_path:
            px = self._icon_cache.get(pt.custom_icon_path)
            if px is None:
                raw = QPixmap(pt.custom_icon_path)
                px  = (
                    raw.scaled(r * 2, r * 2,
                               Qt.AspectRatioMode.KeepAspectRatio,
                               Qt.TransformationMode.SmoothTransformation)
                    if not raw.isNull() else None
                )
                self._icon_cache[pt.custom_icon_path] = px
            if px and not px.isNull():
                painter.drawPixmap(cx - px.width() // 2, cy - px.height() // 2, px)
            else:
                pt.icon_type = "circle"  # fallback при ошибке загрузки

        outline = QPen(QColor("#1e1e2e"), 1.5)
        brush   = QBrush(color)

        if pt.icon_type == "circle":
            painter.setPen(outline)
            painter.setBrush(brush)
            painter.drawEllipse(cx - r, cy - r, 2 * r, 2 * r)

        elif pt.icon_type == "square":
            painter.setPen(outline)
            painter.setBrush(brush)
            painter.drawRect(cx - r, cy - r, 2 * r, 2 * r)

        elif pt.icon_type == "diamond":
            poly = QPolygonF([
                QPointF(cx, cy - r), QPointF(cx + r, cy),
                QPointF(cx, cy + r), QPointF(cx - r, cy),
            ])
            painter.setPen(outline)
            painter.setBrush(brush)
            painter.drawPolygon(poly)

        elif pt.icon_type == "triangle":
            poly = QPolygonF([
                QPointF(cx,     cy - r),
                QPointF(cx + r, cy + r),
                QPointF(cx - r, cy + r),
            ])
            painter.setPen(outline)
            painter.setBrush(brush)
            painter.drawPolygon(poly)

        elif pt.icon_type == "star":
            star = _make_star_points(cx, cy, r, r * 0.42)
            painter.setPen(outline)
            painter.setBrush(brush)
            painter.drawPolygon(star)

        elif pt.icon_type == "pin":
            head_r  = max(3, int(r * 0.65))
            head_cy = cy - r + head_r
            path    = QPainterPath()
            path.addEllipse(QPointF(float(cx), float(head_cy)), head_r, head_r)
            path.moveTo(cx - head_r * 0.65, head_cy + head_r * 0.55)
            path.lineTo(cx + head_r * 0.65, head_cy + head_r * 0.55)
            path.lineTo(cx, cy + r * 0.3)
            path.closeSubpath()
            painter.fillPath(path, brush)
            painter.setPen(outline)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawPath(path)

        elif pt.icon_type == "cross":
            t = max(2, r // 3)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(brush)
            painter.drawRect(cx - r, cy - t, 2 * r, 2 * t)
            painter.drawRect(cx - t, cy - r, 2 * t, 2 * r)
            painter.setPen(outline)
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(cx - r, cy - t, 2 * r, 2 * t)
            painter.drawRect(cx - t, cy - r, 2 * t, 2 * r)

        # Выделение / наведение
        if is_sel:
            painter.setPen(QPen(QColor("#cba6f7"), 2))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawEllipse(cx - r - 5, cy - r - 5, 2 * (r + 5), 2 * (r + 5))
        elif is_hov:
            painter.setPen(QPen(QColor("#f9e2af"), 1, Qt.PenStyle.DashLine))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawEllipse(cx - r - 3, cy - r - 3, 2 * (r + 3), 2 * (r + 3))

    def _draw_temp_measurements(self, painter: QPainter):
        """Рисует временные точки и линии измерения расстояний."""
        pts = self.temp_points
        if not pts:
            return

        W, H = self.width(), self.height()
        font_lbl = QFont("Segoe UI", 8)
        fm_lbl   = QFontMetrics(font_lbl)

        # Линии и метки расстояний между точками
        if len(pts) > 1:
            painter.setOpacity(0.9)
            painter.setPen(QPen(QColor("#f9e2af"), 2, Qt.PenStyle.SolidLine))

            for i in range(len(pts) - 1):
                lat1, lon1 = pts[i]
                lat2, lon2 = pts[i + 1]
                sx1, sy1   = self._lat_lon_to_screen(lat1, lon1)
                sx2, sy2   = self._lat_lon_to_screen(lat2, lon2)

                painter.setPen(QPen(QColor("#f9e2af"), 2))
                painter.setOpacity(0.85)
                painter.drawLine(sx1, sy1, sx2, sy2)

                # Метка расстояния посередине отрезка
                dist = haversine_distance(lat1, lon1, lat2, lon2)
                if dist >= 1000:
                    txt = f"{dist / 1000:.2f} км"
                else:
                    txt = f"{dist:.1f} м"

                mx = (sx1 + sx2) // 2
                my = (sy1 + sy2) // 2
                tw = fm_lbl.horizontalAdvance(txt) + 6
                th = fm_lbl.height()

                painter.setOpacity(0.65)
                painter.fillRect(mx, my - th + 2, tw, th + 2, QColor("#1e1e2e"))
                painter.setOpacity(1.0)
                painter.setFont(font_lbl)
                painter.setPen(QColor("#f9e2af"))
                painter.drawText(mx + 3, my, txt)

        # Маркеры временных точек
        painter.setFont(QFont("Segoe UI", 7, QFont.Weight.Bold))
        for i, (lat, lon) in enumerate(pts):
            sx, sy = self._lat_lon_to_screen(lat, lon)
            if sx < -20 or sy < -20 or sx > W + 20 or sy > H + 20:
                continue

            r = 6
            painter.setOpacity(1.0)
            painter.setPen(QPen(QColor("#1e1e2e"), 2))
            painter.setBrush(QBrush(QColor("#f9e2af")))
            painter.drawEllipse(sx - r, sy - r, r * 2, r * 2)
            painter.setPen(QColor("#1e1e2e"))
            painter.drawText(sx - 3, sy + 3, str(i + 1))

    def _draw_crosshair(self, painter: QPainter, W: int, H: int):
        cx, cy = W // 2, H // 2
        painter.setOpacity(1.0)
        painter.setPen(QPen(QColor("#89b4fa"), 1))
        painter.drawLine(cx - 12, cy, cx + 12, cy)
        painter.drawLine(cx, cy - 12, cx, cy + 12)
        painter.setPen(QPen(QColor("#89b4fa"), 1, Qt.PenStyle.DotLine))
        painter.drawEllipse(cx - 6, cy - 6, 12, 12)

    def _draw_scale(self, painter: QPainter, W: int, H: int):
        """Шкала масштаба + zoom-метка с полупрозрачным фоном."""
        meters_per_pixel = (
            156543.03 * math.cos(math.radians(self.center_lat)) / (2 ** self.zoom)
        )
        scale_m = meters_per_pixel * 100  # метров на 100 пикселей

        # Всегда рисуем фон и zoom-текст
        font_zoom = QFont("Consolas", 9)
        fm_zoom   = QFontMetrics(font_zoom)
        zoom_text = f"zoom {self.zoom}"

        show_scale = scale_m >= 0.05
        px_len     = 0
        label      = ""
        nice       = 0.0

        if show_scale:
            exp  = 10 ** math.floor(math.log10(scale_m))
            nice = round(scale_m / exp) * exp
            px_len = int(nice / meters_per_pixel)
            if px_len < 10 or px_len > W // 2:
                show_scale = False
            else:
                if nice < 1:
                    label = f"{int(round(nice * 100))} см"
                elif nice < 1000:
                    label = f"{int(nice)} м"
                else:
                    label = f"{nice / 1000:.1f} км"

        # Рассчитываем размер фона
        font_scale = QFont("Segoe UI", 8)
        fm_scale   = QFontMetrics(font_scale)

        if show_scale:
            scale_label_w = fm_scale.horizontalAdvance(label)
            bg_w = max(px_len + scale_label_w + 20, fm_zoom.horizontalAdvance(zoom_text) + 16)
        else:
            bg_w = fm_zoom.horizontalAdvance(zoom_text) + 16

        bg_h = 42 if show_scale else 22
        bg_x = 4
        bg_y = H - bg_h - 4

        # Полупрозрачный тёмный фон
        painter.setOpacity(0.55)
        painter.fillRect(bg_x, bg_y, bg_w, bg_h, QColor("#000000"))
        painter.setOpacity(1.0)

        if show_scale:
            x1    = 10
            x2    = x1 + px_len
            bar_y = H - 26

            # Шкала
            painter.setPen(QPen(QColor("#cdd6f4"), 2))
            painter.drawLine(x1, bar_y, x2, bar_y)
            painter.drawLine(x1, bar_y - 4, x1, bar_y + 4)
            painter.drawLine(x2, bar_y - 4, x2, bar_y + 4)

            painter.setFont(font_scale)
            painter.setPen(QColor("#cdd6f4"))
            painter.drawText(x1, bar_y - 6, label)

        # Zoom-метка
        painter.setFont(font_zoom)
        painter.setPen(QColor("#a6adc8"))
        painter.drawText(10, H - 8, zoom_text)

    # ── События мыши ─────────────────────────────────────────────

    def wheelEvent(self, event: QWheelEvent):
        delta    = event.angleDelta().y()
        new_zoom = self.zoom + (1 if delta > 0 else -1)

        if self.layers:
            all_p    = [l.provider for l in self.layers]
            min_z    = min(p.min_zoom for p in all_p)
            max_z    = max(p.max_zoom for p in all_p)
            new_zoom = max(min_z, min(max_z + MAX_OVERZOOM, new_zoom))

        if new_zoom != self.zoom:
            mx, my  = event.position().x(), event.position().y()
            lat_c, lon_c = self._screen_to_lat_lon(mx, my)
            self.zoom    = new_zoom
            tx, ty   = lat_lon_to_tile_f(lat_c, lon_c, self.zoom)
            W, H     = self.width(), self.height()
            cx       = tx - (mx - W / 2.0) / TILE_PX
            cy_      = ty - (my - H / 2.0) / TILE_PX
            self.center_lat, self.center_lon = tile_f_to_lat_lon(cx, cy_, self.zoom)
            self.update()
        event.accept()

    def mousePressEvent(self, event: QMouseEvent):
        mx, my = event.position().x(), event.position().y()
        mods   = QApplication.keyboardModifiers()
        shift  = bool(mods & Qt.KeyboardModifier.ShiftModifier)
        ctrl   = bool(mods & Qt.KeyboardModifier.ControlModifier)

        # ── Средняя кнопка: всегда панорамирование ────────────────
        if event.button() == Qt.MouseButton.MiddleButton:
            self._drag_pos    = event.position()
            self._drag_center = (self.center_lat, self.center_lon)
            return

        if self.point_mode:
            if event.button() == Qt.MouseButton.LeftButton:
                # ─ Ctrl+ЛКМ на динамическую точку → подтвердить ──
                if ctrl:
                    dp_hit = self._find_dynamic_point_at(mx, my)
                    if dp_hit is not None:
                        self._confirm_dynamic_point(dp_hit)
                        return

                hit = self._find_point_at(mx, my)

                if hit is None and not shift and not ctrl:
                    poly_hit = self._find_polygon_at(mx, my)
                    if poly_hit is not None:
                        self._select_polygon(poly_hit)
                        self._push_undo()
                        self._dragging_polygon = poly_hit
                        self._drag_was_moved = False
                        self._drag_start_screen = event.position()
                        self._group_drag_origin = {
                            pid: (pt.lat, pt.lon)
                            for pid in poly_hit.point_ids
                            if (pt := self._get_point_by_id(pid)) is not None
                        }
                        self.update()
                        return

                if hit:
                    alt = bool(mods & Qt.KeyboardModifier.AltModifier)

                    # ─ Ctrl+Alt+ЛКМ: замкнуть цепь на начальную точку
                    if ctrl and alt and self.auto_connect:
                        if (self._chain_start_id
                                and self._last_placed_point_id
                                and hit.id == self._chain_start_id
                                and hit.id != self._last_placed_point_id):
                            self._push_undo()
                            self._create_edge(self._last_placed_point_id, hit.id)
                            self._last_placed_point_id = None
                            self._chain_start_id       = None
                            self.update()
                        return

                    # ─ Alt+ЛКМ: переключить опорную точку ─────────
                    if alt and not ctrl:
                        self._pivot_point_id = (
                            None if self._pivot_point_id == hit.id else hit.id
                        )
                        self.update()
                        return

                    # ─ Режим ручного соединения ───────────────────
                    if self._connect_mode:
                        if self._connect_source_id is None:
                            self._connect_source_id = hit.id
                        elif self._connect_source_id != hit.id:
                            self._push_undo()
                            self._create_edge(self._connect_source_id, hit.id)
                            self._connect_source_id = None
                        self.update()
                        return

                    # ─ Shift+клик: toggle выделения ───────────────
                    if shift:
                        if hit.id in self.selected_point_ids:
                            self.selected_point_ids.discard(hit.id)
                        else:
                            self.selected_point_ids.add(hit.id)
                        self.selection_changed.emit(self._get_selected_points())
                        self.update()
                        return

                    # ─ Клик по невыделенной: заменяем выделение ───
                    if hit.id not in self.selected_point_ids:
                        self.selected_point_ids = {hit.id}
                        self.selection_changed.emit(self._get_selected_points())

                    # ─ Начинаем групповое перетаскивание ──────────
                    self._push_undo()        # снимок перед драгом
                    self._dragging_point    = hit
                    self._drag_was_moved    = False
                    self._drag_start_screen = event.position()
                    self._group_drag_origin = {
                        p.id: (p.lat, p.lon)
                        for p in self.map_points
                        if p.id in self.selected_point_ids
                    }
                    self.update()

                else:
                    # ─ Клик по пустому: снимаем выделение + rubber band
                    if not shift and self.selected_point_ids:
                        self.selected_point_ids.clear()
                        self.selection_changed.emit([])
                    self._rubber_band_start  = event.position()
                    self._rubber_band_end    = event.position()
                    self._rubber_band_active = False
                    self.update()

            elif event.button() == Qt.MouseButton.RightButton:
                hit = self._find_point_at(mx, my)
                if hit:
                    if ctrl:
                        # Ctrl+ПКМ: удаляем точку (или всё выделение)
                        if hit.id in self.selected_point_ids:
                            for p in self._get_selected_points():
                                self._delete_map_point(p)
                        else:
                            self._delete_map_point(hit)
                    else:
                        self._show_point_ctx_menu(hit, event.globalPosition().toPoint())
            return   # ← не передаём событие дальше в не-point режим

        # ── Не-point режимы: ЛКМ = начало панорамирования ─────────
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_pos    = event.position()
            self._drag_center = (self.center_lat, self.center_lon)
        elif event.button() == Qt.MouseButton.RightButton and self.view_mode:
            if self.temp_points:
                self.temp_points.pop()
                self.temp_pts_changed.emit()
                self.update()

    def mouseMoveEvent(self, event: QMouseEvent):
        mx, my = event.position().x(), event.position().y()
        lat, lon = self._screen_to_lat_lon(mx, my)
        self.coord_hovered.emit(lat, lon)

        # ── R-поворот мышью ───────────────────────────────────────
        if (self._rotate_mode
                and self._rotate_screen_ref is not None
                and self._rotate_center_screen is not None
                and self._rotate_center_latlon is not None):
            cx_s, cy_s   = self._rotate_center_screen
            ref_x, ref_y = self._rotate_screen_ref
            angle_ref = math.atan2(ref_y - cy_s, ref_x - cx_s)
            angle_cur = math.atan2(my   - cy_s, mx    - cx_s)
            delta     = angle_ref - angle_cur
            cos_a   = math.cos(delta)
            sin_a   = math.sin(delta)
            c_lat, c_lon = self._rotate_center_latlon
            cos_lat = math.cos(math.radians(c_lat))
            for pt in self.map_points:
                if pt.id in self._rotate_origins and pt.id != self._pivot_point_id:
                    o_lat, o_lon = self._rotate_origins[pt.id]
                    dx =  (o_lon - c_lon) * cos_lat
                    dy =   o_lat - c_lat
                    pt.lon = c_lon + (dx * cos_a - dy * sin_a) / cos_lat
                    pt.lat = c_lat + (dx * sin_a + dy * cos_a)
            self._update_timer.start(16)
            return

        # ── Групповое перетаскивание точек ────────────────────────
        if (self._dragging_point is not None or self._dragging_polygon is not None) and self._drag_start_screen is not None:
            dx = mx - self._drag_start_screen.x()
            dy = my - self._drag_start_screen.y()
            if abs(dx) > 2 or abs(dy) > 2 or self._drag_was_moved:
                self._drag_was_moved = True
                for pt in self.map_points:
                    if pt.id in self._group_drag_origin:
                        o_lat, o_lon = self._group_drag_origin[pt.id]
                        o_sx, o_sy   = self._lat_lon_to_screen(o_lat, o_lon)
                        pt.lat, pt.lon = self._screen_to_lat_lon(o_sx + dx, o_sy + dy)
                self._update_timer.start(16)
            return

        # ── Rubber band (рамка выделения) ─────────────────────────
        if self._rubber_band_start is not None:
            self._rubber_band_end = event.position()
            dx = abs(mx - self._rubber_band_start.x())
            dy = abs(my - self._rubber_band_start.y())
            if dx > 5 or dy > 5:
                self._rubber_band_active = True
            self._update_timer.start(16)
            return

        # ── Hover-эффект в режиме точек ───────────────────────────
        if self.point_mode:
            hit     = self._find_point_at(mx, my)
            new_hov = hit.id if hit else None
            if new_hov != self._hovered_point_id:
                self._hovered_point_id = new_hov
                self._update_timer.start(16)
            # Hover для рёбер (только когда нет точки под курсором)
            new_ehov = None
            if not hit and self.map_edges:
                e_hit    = self._find_edge_at(mx, my)
                new_ehov = e_hit.id if e_hit else None
            if new_ehov != self._hovered_edge_id:
                self._hovered_edge_id = new_ehov
                self._update_timer.start(16)

        # ── Панорамирование (ЛКМ или СКМ) ────────────────────────
        if self._drag_pos is not None:
            dx = mx - self._drag_pos.x()
            dy = my - self._drag_pos.y()
            if abs(dx) > 1 or abs(dy) > 1:
                if self._drag_center:
                    cx, cy = lat_lon_to_tile_f(
                        self._drag_center[0], self._drag_center[1], self.zoom
                    )
                    W, H   = self.width(), self.height()
                    new_cx = cx - dx / TILE_PX
                    new_cy = cy - dy / TILE_PX
                    self.center_lat, self.center_lon = tile_f_to_lat_lon(
                        new_cx, new_cy, self.zoom
                    )
                    self._update_timer.start(16)

    def mouseReleaseEvent(self, event: QMouseEvent):
        # ── Средняя кнопка: завершаем панорамирование ─────────────
        if event.button() == Qt.MouseButton.MiddleButton:
            self._drag_pos    = None
            self._drag_center = None
            return

        if event.button() == Qt.MouseButton.LeftButton:
            # ── Завершаем групповое перетаскивание ────────────────
            if self._dragging_point is not None or self._dragging_polygon is not None:
                if self._drag_was_moved:
                    for pt in self._get_selected_points():
                        self.point_moved.emit(pt)
                    self._check_interior_subdivisions()
                    self._deduplicate_polygons()
                    self._sync_dynamic_points()
                self._dragging_point    = None
                self._dragging_polygon  = None
                self._drag_was_moved    = False
                self._drag_start_screen = None
                self._group_drag_origin = {}
                self.update()
                return

            # ── Завершаем rubber band ─────────────────────────────
            if self._rubber_band_start is not None:
                if self._rubber_band_active and self._rubber_band_end is not None:
                    shift = bool(
                        QApplication.keyboardModifiers() & Qt.KeyboardModifier.ShiftModifier
                    )
                    x1 = min(self._rubber_band_start.x(), self._rubber_band_end.x())
                    y1 = min(self._rubber_band_start.y(), self._rubber_band_end.y())
                    x2 = max(self._rubber_band_start.x(), self._rubber_band_end.x())
                    y2 = max(self._rubber_band_start.y(), self._rubber_band_end.y())
                    if not shift:
                        self.selected_point_ids.clear()
                    for pt in self.map_points:
                        sx, sy = self._lat_lon_to_screen(pt.lat, pt.lon)
                        if x1 <= sx <= x2 and y1 <= sy <= y2:
                            self.selected_point_ids.add(pt.id)
                    self.selection_changed.emit(self._get_selected_points())
                else:
                    # Клик без движения в точечном режиме
                    if self.point_mode:
                        mods_r = QApplication.keyboardModifiers()
                        ctrl_r = bool(mods_r & Qt.KeyboardModifier.ControlModifier)
                        if ctrl_r:
                            # Ctrl+ЛКМ — создать точку
                            lat, lon = self._screen_to_lat_lon(
                                self._rubber_band_start.x(), self._rubber_band_start.y()
                            )
                            self._add_map_point_at(lat, lon)
                        else:
                            # Простой ЛКМ по пустому месту — снять выделение
                            if self.selected_point_ids:
                                self.selected_point_ids.clear()
                                self.selection_changed.emit([])
                self._rubber_band_start  = None
                self._rubber_band_end    = None
                self._rubber_band_active = False
                self.update()
                return

            # ── ЛКМ в не-point режимах ────────────────────────────
            if self._drag_pos is not None:
                dx = abs(event.position().x() - self._drag_pos.x())
                dy = abs(event.position().y() - self._drag_pos.y())
                if dx < 5 and dy < 5:
                    lat, lon = self._screen_to_lat_lon(
                        event.position().x(), event.position().y()
                    )
                    if self.view_mode:
                        self.temp_points.append((lat, lon))
                        self.temp_pts_changed.emit()
                        self.update()
                    elif not self.point_mode:
                        self.coord_clicked.emit(lat, lon)
            self._drag_pos    = None
            self._drag_center = None

    def keyPressEvent(self, event):
        if event.isAutoRepeat():
            super().keyPressEvent(event)
            return

        key  = event.key()
        mods = event.modifiers()
        ctrl  = bool(mods & Qt.KeyboardModifier.ControlModifier)
        shift = bool(mods & Qt.KeyboardModifier.ShiftModifier)

        # Ctrl+Z — Undo
        if key == Qt.Key.Key_Z and ctrl and not shift:
            self.undo()
            return

        # Ctrl+Shift+Z — Redo
        if key == Qt.Key.Key_Z and ctrl and shift:
            self.redo()
            return

        # H — переключить глобальные метки
        if key == Qt.Key.Key_H and not ctrl:
            self.show_geo_labels = not self.show_geo_labels
            on = self.show_geo_labels
            self.show_point_labels = on
            self.show_coord_labels = on
            self.show_angle_labels = on
            self.show_area_labels = on
            self.show_edge_length_labels = on
            self.labels_toggled.emit(self.show_geo_labels)
            self.update()
            return

        # R — начать поворот мышью (только в режиме точек)
        if key == Qt.Key.Key_R and not ctrl and self.point_mode:
            pts = self._get_selected_points()
            if len(pts) >= 2 or (len(pts) >= 1 and self._pivot_point_id):
                c = self._get_selection_center()
                if self._pivot_point_id:
                    piv = self._get_point_by_id(self._pivot_point_id)
                    if piv:
                        c = (piv.lat, piv.lon)
                if c:
                    self._push_undo()
                    self._rotate_center_latlon = c
                    cx_s, cy_s = self._lat_lon_to_screen(c[0], c[1])
                    self._rotate_center_screen = (cx_s, cy_s)
                    from PyQt6.QtGui import QCursor
                    local = self.mapFromGlobal(QCursor.pos())
                    self._rotate_screen_ref = (local.x(), local.y())
                    self._rotate_origins = {
                        p.id: (p.lat, p.lon) for p in pts
                    }
                    self._rotate_mode = True
                    self.update()
            return

        # Escape — сброс в точечном режиме
        if key == Qt.Key.Key_Escape and self.point_mode:
            if self._rotate_mode:
                self._rotate_mode = False
                # Восстанавливаем позиции из origins
                for pt in self.map_points:
                    if pt.id in self._rotate_origins:
                        pt.lat, pt.lon = self._rotate_origins[pt.id]
                self._rotate_origins = {}
                self._undo_stack.pop() if self._undo_stack else None
            elif self._connect_mode:
                self._connect_source_id = None
            else:
                self.selected_point_ids.clear()
                self._pivot_point_id = None
                self.selection_changed.emit([])
            self.update()
            return

        super().keyPressEvent(event)

    def keyReleaseEvent(self, event):
        if event.isAutoRepeat():
            super().keyReleaseEvent(event)
            return
        if event.key() == Qt.Key.Key_R and self._rotate_mode:
            self._rotate_mode    = False
            self._rotate_origins = {}
            self._rotate_screen_ref    = None
            self._rotate_center_screen = None
            self._rotate_center_latlon = None
            self.update()
        super().keyReleaseEvent(event)

    # ── Вспомогательные методы системы точек ─────────────────────

    def _find_edge_at(self, sx: float, sy: float, threshold: int = 8) -> Optional["MapEdge"]:
        """Ищет ближайшее ребро по расстоянию от курсора до отрезка."""
        best_e = None
        best_d = float("inf")
        for e in self.map_edges:
            a = self._get_point_by_id(e.point_a_id)
            b = self._get_point_by_id(e.point_b_id)
            if a is None or b is None:
                continue
            ax, ay = self._lat_lon_to_screen(a.lat, a.lon)
            bx, by = self._lat_lon_to_screen(b.lat, b.lon)
            dx, dy = bx - ax, by - ay
            if dx == 0 and dy == 0:
                dist = math.hypot(sx - ax, sy - ay)
            else:
                t    = max(0.0, min(1.0, ((sx - ax)*dx + (sy - ay)*dy) / (dx*dx + dy*dy)))
                dist = math.hypot(sx - (ax + t*dx), sy - (ay + t*dy))
            if dist <= threshold and dist < best_d:
                best_d = dist
                best_e = e
        return best_e

    def _find_point_at(self, sx: float, sy: float, radius: int = 14) -> Optional[MapPoint]:
        """Ищет ближайшую точку в заданном радиусе от экранной координаты."""
        best_pt   = None
        best_dist = float("inf")
        for pt in self.map_points:
            px, py = self._lat_lon_to_screen(pt.lat, pt.lon)
            dist   = math.hypot(sx - px, sy - py)
            if dist <= radius and dist < best_dist:
                best_dist = dist
                best_pt   = pt
        return best_pt

    # ── Базовые геттеры ───────────────────────────────────────────

    def _get_point_by_id(self, point_id: str) -> Optional[MapPoint]:
        for pt in self.map_points:
            if pt.id == point_id:
                return pt
        return None

    def _get_selected_points(self) -> List[MapPoint]:
        return [pt for pt in self.map_points if pt.id in self.selected_point_ids]

    def _get_selection_center(self) -> Optional[Tuple[float, float]]:
        pts = self._get_selected_points()
        if not pts:
            return None
        return sum(p.lat for p in pts) / len(pts), sum(p.lon for p in pts) / len(pts)

    # ── Вращение выделения ────────────────────────────────────────

    def rotate_selection(self, angle_deg: float):
        """Вращает выбранные точки вокруг центра (или pivot-точки). Кнопки MapTab."""
        pts = self._get_selected_points()
        if len(pts) < 2:
            return
        self._push_undo()
        if self._pivot_point_id:
            pivot = self._get_point_by_id(self._pivot_point_id)
            clat, clon = (pivot.lat, pivot.lon) if pivot else self._get_selection_center() or (0.0, 0.0)
        else:
            c = self._get_selection_center()
            if c is None:
                return
            clat, clon = c
        rad     = math.radians(angle_deg)
        cos_a   = math.cos(rad)
        sin_a   = math.sin(rad)
        cos_lat = math.cos(math.radians(clat))
        for pt in pts:
            if pt.id == self._pivot_point_id:
                continue
            dx =  (pt.lon - clon) * cos_lat
            dy =   pt.lat - clat
            pt.lon = clon + (dx * cos_a - dy * sin_a) / cos_lat
            pt.lat = clat + (dx * sin_a + dy * cos_a)
        self.update()

    # ── Геометрические вычисления ─────────────────────────────────

    def _calc_bearing(self, lat1: float, lon1: float, lat2: float, lon2: float) -> float:
        """Азимут (0–360°) от точки 1 к точке 2."""
        d_lon  = math.radians(lon2 - lon1)
        r_lat1 = math.radians(lat1)
        r_lat2 = math.radians(lat2)
        x = math.sin(d_lon) * math.cos(r_lat2)
        y = math.cos(r_lat1) * math.sin(r_lat2) - math.sin(r_lat1) * math.cos(r_lat2) * math.cos(d_lon)
        return (math.degrees(math.atan2(x, y)) + 360) % 360

    def _calc_edge_length(self, edge: "MapEdge") -> float:
        """Длина ребра в метрах (Haversine)."""
        a = self._get_point_by_id(edge.point_a_id)
        b = self._get_point_by_id(edge.point_b_id)
        if a is None or b is None:
            return 0.0
        return haversine_distance(a.lat, a.lon, b.lat, b.lon)

    def _calc_polygon_area(self, poly: "MapPolygon") -> float:
        """Площадь полигона в кв. метрах (формула Гаусса + экв. проекция)."""
        pts = [self._get_point_by_id(pid) for pid in poly.point_ids]
        pts = [p for p in pts if p is not None]
        n   = len(pts)
        if n < 3:
            return 0.0
        clat    = sum(p.lat for p in pts) / n
        R       = 6_371_000.0
        cos_lat = math.cos(math.radians(clat))
        xs = [math.radians(p.lon) * R * cos_lat for p in pts]
        ys = [math.radians(p.lat) * R            for p in pts]
        area = 0.0
        for i in range(n):
            j     = (i + 1) % n
            area += xs[i] * ys[j] - xs[j] * ys[i]
        return abs(area) / 2.0

    def _polygon_vertices_as_xy(
        self,
        poly: "MapPolygon",
        ref_lat: Optional[float] = None,
    ) -> List[Tuple[float, float]]:
        """Возвращает вершины полигона в метрах (эквирект. проекция с общим ref_lat)."""
        pts = [self._get_point_by_id(pid) for pid in poly.point_ids]
        pts = [p for p in pts if p is not None]
        if len(pts) < 3:
            return []
        clat = ref_lat if ref_lat is not None else sum(p.lat for p in pts) / len(pts)
        R = 6_371_000.0
        cos_lat = math.cos(math.radians(clat))
        return [
            (math.radians(p.lon) * R * cos_lat, math.radians(p.lat) * R)
            for p in pts
        ]

    def _effective_polygon_area(self, poly: "MapPolygon") -> float:
        """Площадь полигона без перекрытий (динамическое деление для любых фигур)."""
        if ShapelyPolygon is None or unary_union is None:
            return self._calc_polygon_area(poly)

        target_pts = [self._get_point_by_id(pid) for pid in poly.point_ids]
        target_pts = [p for p in target_pts if p is not None]
        if len(target_pts) < 3:
            return 0.0

        ref_lat = sum(p.lat for p in target_pts) / len(target_pts)
        target_vertices = self._polygon_vertices_as_xy(poly, ref_lat=ref_lat)
        if len(target_vertices) < 3:
            return 0.0
        target_shape = ShapelyPolygon(target_vertices)
        if target_shape.is_empty or not target_shape.is_valid:
            return max(0.0, target_shape.buffer(0).area if not target_shape.is_empty else 0.0)

        try:
            target_idx = self.map_polygons.index(poly)
        except ValueError:
            target_idx = -1

        subtract_shapes = []
        for other in self.map_polygons[target_idx + 1:]:
            verts = self._polygon_vertices_as_xy(other, ref_lat=ref_lat)
            if len(verts) < 3:
                continue
            try:
                shp = ShapelyPolygon(verts)
            except Exception:
                continue
            if shp.is_empty:
                continue
            if not shp.is_valid:
                shp = shp.buffer(0)
            if shp.is_empty:
                continue
            subtract_shapes.append(shp)

        if not subtract_shapes:
            return target_shape.area

        try:
            subtraction_union = unary_union(subtract_shapes)
            effective_shape = target_shape.difference(subtraction_union)
        except Exception:
            return target_shape.area

        if effective_shape.is_empty:
            return 0.0

        if not effective_shape.is_valid:
            effective_shape = effective_shape.buffer(0)
            if effective_shape.is_empty:
                return 0.0

        return max(0.0, float(effective_shape.area))

    def _calc_total_effective_area(self) -> float:
        return sum(self._effective_polygon_area(poly) for poly in self.map_polygons)

    # ── Работа с рёбрами ──────────────────────────────────────────

    def _find_path(self, adj: Dict[str, set], start: str, end: str) -> Optional[List[str]]:
        """BFS: ищет кратчайший путь от start до end в графе adj."""
        queue   = deque([[start]])
        visited = {start}
        while queue:
            path = queue.popleft()
            node = path[-1]
            for nb in adj.get(node, set()):
                if nb == end:
                    return path + [nb]
                if nb not in visited:
                    visited.add(nb)
                    queue.append(path + [nb])
        return None

    def _auto_detect_polygon(self, new_edge: "MapEdge") -> Optional["MapPolygon"]:
        """Пытается обнаружить замкнутый цикл после добавления ребра."""
        adj: Dict[str, set] = {}
        for e in self.map_edges:
            adj.setdefault(e.point_a_id, set()).add(e.point_b_id)
            adj.setdefault(e.point_b_id, set()).add(e.point_a_id)
        a_id = new_edge.point_a_id
        b_id = new_edge.point_b_id
        # Строим граф без самого нового ребра и ищем обходной путь b→a
        adj_no = {
            k: (v - {a_id} if k == b_id else v - {b_id} if k == a_id else set(v))
            for k, v in adj.items()
        }
        path = self._find_path(adj_no, b_id, a_id)
        if path is None or len(path) < 3:
            return None
        poly_set = set(path)
        for existing in self.map_polygons:
            if set(existing.point_ids) == poly_set:
                return None
        return MapPolygon(point_ids=path)

    def _create_edge(self, id_a: str, id_b: str) -> Optional["MapEdge"]:
        """Создаёт ребро. Если оба конца в одном полигоне — авто-разрезает его."""
        if id_a == id_b:
            return None
        for e in self.map_edges:
            if {e.point_a_id, e.point_b_id} == {id_a, id_b}:
                return None  # уже существует

        edge = MapEdge(
            point_a_id = id_a,
            point_b_id = id_b,
            color      = self.default_edge_color,
            width      = self.default_edge_width,
            line_type  = self.default_edge_type,
            opacity    = self.default_edge_opacity,
        )
        self.map_edges.append(edge)
        self.edge_added.emit(edge)

        # Авто-разрез: оба конца ребра в одном полигоне → сплит
        split_done = False
        for poly in list(self.map_polygons):
            if id_a in poly.point_ids and id_b in poly.point_ids:
                self._split_polygon_only(poly, id_a, id_b)
                split_done = True
                break

        # Авто-обнаружение пересечений с другими рёбрами → DynamicPoint
        has_intersection = False
        if self.auto_dynamic_points:
            has_intersection = self._check_new_intersections(edge)

        # Авто-полигон по замкнутому циклу (только если разреза не было и нет пересечений)
        if not split_done and not has_intersection and self.auto_polygon:
            new_poly = self._auto_detect_polygon(edge)
            if new_poly is not None:
                self.map_polygons.append(new_poly)

        return edge

    def connect_selected_points(self):
        """Соединяет рёбрами все выбранные точки по порядку.
        Если среди выбранных есть внутренняя точка полигона — разбивает полигон на секторы."""
        pts = self._get_selected_points()
        if len(pts) < 2:
            return
        self._push_undo()
        for i in range(len(pts) - 1):
            self._create_edge(pts[i].id, pts[i + 1].id)
        # Проверяем, не появились ли внутренние точки, соединённые с вершинами полигона
        self._check_interior_subdivisions()
        self.update()

    def delete_edges_for_selection(self):
        """Удаляет рёбра, у которых хотя бы один конец выбран."""
        self._push_undo()
        to_remove = [
            e for e in self.map_edges
            if e.point_a_id in self.selected_point_ids
            or e.point_b_id in self.selected_point_ids
        ]
        for e in to_remove:
            self.map_edges.remove(e)
            self.edge_deleted.emit(e)

        if to_remove:
            self._merge_polygons_by_removed_edges(to_remove)

        pt_ids = {p.id for p in self.map_points}
        edge_pairs = (
            {(e.point_a_id, e.point_b_id) for e in self.map_edges}
            | {(e.point_b_id, e.point_a_id) for e in self.map_edges}
        )
        self.map_polygons = [
            poly for poly in self.map_polygons
            if all(pid in pt_ids for pid in poly.point_ids)
            and all(
                (poly.point_ids[i], poly.point_ids[(i + 1) % len(poly.point_ids)]) in edge_pairs
                for i in range(len(poly.point_ids))
            )
        ]
        self._deduplicate_polygons()
        self._sync_dynamic_points()
        self.update()

    # ── Работа с полигонами ───────────────────────────────────────

    def create_polygon_from_selected(self):
        """Создаёт полигон из всех выбранных точек (по порядку)."""
        pts = self._get_selected_points()
        if len(pts) < 3:
            return
        self._push_undo()
        poly_ids = [p.id for p in pts]
        for i in range(len(poly_ids)):
            self._create_edge_raw(poly_ids[i], poly_ids[(i + 1) % len(poly_ids)])
        poly = MapPolygon(
            point_ids=poly_ids,
            fill_color=self.default_poly_fill,
            fill_opacity=self.default_poly_fill_opacity,
            border_color=self.default_poly_border,
            border_width=self.default_poly_border_width,
            border_type=self.default_poly_border_type,
        )
        self.map_polygons.append(poly)
        self._selected_polygon_id = poly.id
        self._sync_dynamic_points()
        self.update()

    def import_polygon_points(self, lat_lon_points: List[Tuple[float, float]], auto_point_names: bool = False) -> bool:
        """Импортирует полигон по списку (lat, lon) и создаёт рёбра по порядку."""
        if len(lat_lon_points) < 3:
            return False
        self._push_undo()
        created_pts: List[MapPoint] = []
        for lat, lon in lat_lon_points:
            idx = len(created_pts) + 1
            pt = MapPoint(
                lat=lat,
                lon=lon,
                label=str(idx) if auto_point_names else "",
                icon_type=self.default_icon_type,
                color=self.default_color,
                size=self.default_size,
            )
            self.map_points.append(pt)
            created_pts.append(pt)
            self.point_added.emit(pt)

        created_ids = [pt.id for pt in created_pts]
        for i in range(len(created_ids)):
            self._create_edge_raw(created_ids[i], created_ids[(i + 1) % len(created_ids)])

        poly = MapPolygon(
            point_ids=created_ids,
            fill_color=self.default_poly_fill,
            fill_opacity=self.default_poly_fill_opacity,
            border_color=self.default_poly_border,
            border_width=self.default_poly_border_width,
            border_type=self.default_poly_border_type,
        )
        self.map_polygons.append(poly)
        self._selected_polygon_id = poly.id
        self.selected_point_ids = set(created_ids)
        self.selection_changed.emit(created_pts)
        self._last_placed_point_id = created_ids[-1]
        self._chain_start_id = created_ids[0]
        self._sync_dynamic_points()
        self.update()
        return True


    def _get_selected_polygon(self) -> Optional["MapPolygon"]:
        if not self._selected_polygon_id:
            return None
        return next((p for p in self.map_polygons if p.id == self._selected_polygon_id), None)

    def _find_polygon_at(self, sx: float, sy: float) -> Optional["MapPolygon"]:
        for poly in reversed(self.map_polygons):
            pts = [self._get_point_by_id(pid) for pid in poly.point_ids]
            if any(p is None for p in pts) or len(pts) < 3:
                continue
            path = QPainterPath()
            screen = [QPointF(*self._lat_lon_to_screen(p.lat, p.lon)) for p in pts]
            path.moveTo(screen[0])
            for sp in screen[1:]:
                path.lineTo(sp)
            path.closeSubpath()
            if path.contains(QPointF(float(sx), float(sy))):
                return poly
        return None

    def _select_polygon(self, poly: Optional["MapPolygon"]):
        self._selected_polygon_id = poly.id if poly else None
        if poly is None:
            self.selected_point_ids.clear()
            self.selection_changed.emit([])
            return
        self.selected_point_ids = set(poly.point_ids)
        self.selection_changed.emit(self._get_selected_points())

    def _renumber_polygon_points_clockwise(self, poly: "MapPolygon", start_pid: str):
        ids = list(poly.point_ids)
        if start_pid not in ids:
            return
        start_idx = ids.index(start_pid)
        ordered = ids[start_idx:] + ids[:start_idx]
        for idx, pid in enumerate(ordered, start=1):
            pt = self._get_point_by_id(pid)
            if pt:
                pt.label = str(idx)

    def delete_polygon(self, poly: "MapPolygon"):
        """Удаляет полигон."""
        if poly in self.map_polygons:
            self._push_undo()
            self.map_polygons.remove(poly)
            self.update()

    # ── Undo / Redo ───────────────────────────────────────────────

    _UNDO_MAX = 50

    def _snapshot(self) -> dict:
        return {
            "points":   copy.deepcopy(self.map_points),
            "edges":    copy.deepcopy(self.map_edges),
            "polygons": copy.deepcopy(self.map_polygons),
            "dynamic":  copy.deepcopy(self.dynamic_points),
        }

    def _push_undo(self):
        self._undo_stack.append(self._snapshot())
        if len(self._undo_stack) > self._UNDO_MAX:
            self._undo_stack.pop(0)
        self._redo_stack.clear()

    def _restore(self, snap: dict):
        self.map_points    = snap["points"]
        self.map_edges     = snap["edges"]
        self.map_polygons  = snap["polygons"]
        self.dynamic_points = snap.get("dynamic", [])
        self.selected_point_ids.clear()
        self._hovered_point_id  = None
        self._hovered_edge_id   = None
        self._pivot_point_id    = None
        self._connect_source_id    = None
        self._selected_polygon_id  = None
        self._last_placed_point_id = None
        self._chain_start_id       = None
        self.selection_changed.emit([])
        self.update()

    def undo(self):
        if not self._undo_stack:
            return
        self._redo_stack.append(self._snapshot())
        self._restore(self._undo_stack.pop())

    def redo(self):
        if not self._redo_stack:
            return
        self._undo_stack.append(self._snapshot())
        self._restore(self._redo_stack.pop())

    # ── Геометрия: точка в полигоне ──────────────────────────

    def _point_in_polygon_latlon(self, lat: float, lon: float,
                                 poly: "MapPolygon") -> bool:
        """Рай-casting тест: лежит ли (lat, lon) внутри poly (в координатах lat/lon)."""
        pts = [self._get_point_by_id(pid) for pid in poly.point_ids]
        pts = [p for p in pts if p is not None]
        n   = len(pts)
        if n < 3:
            return False
        x, y  = lon, lat
        inside = False
        j      = n - 1
        for i in range(n):
            xi, yi = pts[i].lon, pts[i].lat
            xj, yj = pts[j].lon, pts[j].lat
            if ((yi > y) != (yj > y)) and (
                x < (xj - xi) * (y - yi) / (yj - yi) + xi
            ):
                inside = not inside
            j = i
        return inside

    # ── Разбиение полигона из внутренней точки ────────────────────

    def _subdivide_by_interior(self, poly: "MapPolygon",
                                interior_id: str,
                                connected_vertex_ids: list):
        """Заменяет poly на K секторов: [D, Vi, ..., Vj] для каждой пары соседних
        соединённых вершин полигона."""
        ids = poly.point_ids
        conn_in_poly = [vid for vid in connected_vertex_ids if vid in ids]
        if len(conn_in_poly) < 2:
            return
        sorted_indices = sorted(ids.index(vid) for vid in conn_in_poly)
        k  = len(sorted_indices)
        kw = dict(
            fill_color   = poly.fill_color,   fill_opacity = poly.fill_opacity,
            border_color = poly.border_color, border_width = poly.border_width,
            border_type  = poly.border_type,
        )
        if poly in self.map_polygons:
            self.map_polygons.remove(poly)
        for j in range(k):
            i_start = sorted_indices[j]
            i_end   = sorted_indices[(j + 1) % k]
            if i_end > i_start:
                sector = ids[i_start : i_end + 1]
            else:
                sector = ids[i_start:] + ids[: i_end + 1]
            sub_ids = [interior_id] + sector
            if len(sub_ids) >= 3:
                cand_set = set(sub_ids)
                if any(set(p.point_ids) == cand_set for p in self.map_polygons):
                    continue
                self.map_polygons.append(MapPolygon(point_ids=sub_ids, **kw))

    def _deduplicate_polygons(self):
        """Удаляет дубли полигонов с одинаковым набором вершин."""
        uniq: List[MapPolygon] = []
        seen = set()
        for poly in self.map_polygons:
            key = frozenset(poly.point_ids)
            if len(poly.point_ids) < 3 or key in seen:
                continue
            seen.add(key)
            uniq.append(poly)
        self.map_polygons = uniq

    def _polygon_contains_edge(self, poly: "MapPolygon", a_id: str, b_id: str) -> bool:
        ids = poly.point_ids
        n = len(ids)
        for i in range(n):
            j = (i + 1) % n
            if (ids[i], ids[j]) in ((a_id, b_id), (b_id, a_id)):
                return True
        return False

    def _long_chain_between_adjacent(self, ids: List[str], start_id: str, end_id: str) -> Optional[List[str]]:
        """Возвращает длинную цепочку от start_id до end_id (в обход ребра start-end)."""
        if start_id not in ids or end_id not in ids:
            return None
        n = len(ids)
        if n < 3:
            return None
        i_start = ids.index(start_id)
        i_end = ids.index(end_id)

        fwd = [start_id]
        i = i_start
        while i != i_end:
            i = (i + 1) % n
            fwd.append(ids[i])

        bwd = [start_id]
        i = i_start
        while i != i_end:
            i = (i - 1 + n) % n
            bwd.append(ids[i])

        chain = fwd if len(fwd) > len(bwd) else bwd
        return chain if len(chain) >= 3 else None

    def _merge_polygons_by_removed_edges(self, removed_edges: List["MapEdge"]):
        """Если удалено внутреннее ребро между полигонами — сливает их в один."""
        for edge in removed_edges:
            a_id, b_id = edge.point_a_id, edge.point_b_id
            while True:
                candidates = [
                    p for p in self.map_polygons
                    if self._polygon_contains_edge(p, a_id, b_id)
                ]
                if len(candidates) < 2:
                    break

                p1, p2 = candidates[0], candidates[1]
                chain1 = self._long_chain_between_adjacent(p1.point_ids, a_id, b_id)
                chain2 = self._long_chain_between_adjacent(p2.point_ids, a_id, b_id)
                if not chain1 or not chain2:
                    break

                merged_ids = chain1 + list(reversed(chain2))[1:-1]
                dedup_ids = []
                for pid in merged_ids:
                    if not dedup_ids or dedup_ids[-1] != pid:
                        dedup_ids.append(pid)
                if len(set(dedup_ids)) < 3:
                    break

                kw = dict(
                    fill_color=p1.fill_color,
                    fill_opacity=p1.fill_opacity,
                    border_color=p1.border_color,
                    border_width=p1.border_width,
                    border_type=p1.border_type,
                )
                self.map_polygons.remove(p1)
                self.map_polygons.remove(p2)
                self.map_polygons.append(MapPolygon(point_ids=dedup_ids, **kw))

    def _sync_dynamic_points(self):
        """Синхронизирует динамические точки с текущей геометрией рёбер."""
        if not self.show_dynamic_points:
            self.dynamic_points.clear()
            return
        if not self.auto_dynamic_points:
            self._remove_stale_dynamic_points()
            return
        self.dynamic_points.clear()
        for i, ea in enumerate(self.map_edges):
            for eb in self.map_edges[i + 1:]:
                if {ea.point_a_id, ea.point_b_id} & {eb.point_a_id, eb.point_b_id}:
                    continue
                p1 = self._get_point_by_id(ea.point_a_id)
                p2 = self._get_point_by_id(ea.point_b_id)
                p3 = self._get_point_by_id(eb.point_a_id)
                p4 = self._get_point_by_id(eb.point_b_id)
                if not (p1 and p2 and p3 and p4):
                    continue
                if self._segment_intersection(p1.lat, p1.lon, p2.lat, p2.lon,
                                              p3.lat, p3.lon, p4.lat, p4.lon) is None:
                    continue
                self.dynamic_points.append(DynamicPoint(edge_a_id=ea.id, edge_b_id=eb.id))
    def _check_interior_subdivisions(self):
        """Сканирует все точки: если точка внутри полигона и соединена с 2+ его
        вершинами — разбивает полигон на сектора. Вызывается после connect_selected."""
        if not self.show_dynamic_points:
            return

        # Если точка уже является вершиной какого-либо полигона, не нужно
        # автоматически делить другие полигоны через неё. Иначе при временном
        # заходе «прилегающей» точки внутрь соседней фигуры можно получить
        # накопление площади (большой полигон + отдельный треугольник).
        polygon_vertex_ids = {
            pid
            for poly in self.map_polygons
            for pid in poly.point_ids
        }

        for pt in list(self.map_points):
            if pt.id in polygon_vertex_ids:
                continue
            for poly in list(self.map_polygons):
                if pt.id in poly.point_ids:
                    continue
                if self._point_in_polygon_latlon(pt.lat, pt.lon, poly):
                    conn_vids = [
                        vid for vid in poly.point_ids
                        if any(
                            {e.point_a_id, e.point_b_id} == {pt.id, vid}
                            for e in self.map_edges
                        )
                    ]
                    if len(conn_vids) >= 2:
                        self._subdivide_by_interior(poly, pt.id, conn_vids)
                        self._deduplicate_polygons()
                    break

    def _create_edge_raw(self, id_a: str, id_b: str) -> "Optional[MapEdge]":
        """Создаёт ребро без авто-полигона, авто-разреза и обнаружения пересечений."""
        if id_a == id_b:
            return None
        for e in self.map_edges:
            if {e.point_a_id, e.point_b_id} == {id_a, id_b}:
                return None
        edge = MapEdge(
            point_a_id = id_a, point_b_id = id_b,
            color      = self.default_edge_color,
            width      = self.default_edge_width,
            line_type  = self.default_edge_type,
            opacity    = self.default_edge_opacity,
        )
        self.map_edges.append(edge)
        self.edge_added.emit(edge)
        if self.auto_dynamic_points:
            self._check_new_intersections(edge)
        return edge
    # ── Пересечения рёбер (динамические точки) ────────────────────

    def _segment_intersection(self,
                              p1lat: float, p1lon: float,
                              p2lat: float, p2lon: float,
                              p3lat: float, p3lon: float,
                              p4lat: float, p4lon: float,
                              ) -> "Optional[Tuple[float, float, float, float]]":  # type: ignore
        """Возвращает (t, u, lat, lon) если отрезки [p1,p2] и [p3,p4] пересекаются
        строго внутри обоих (t,u ∈ (eps, 1-eps)), иначе None."""
        dx12 = p2lon - p1lon;  dy12 = p2lat - p1lat
        dx34 = p4lon - p3lon;  dy34 = p4lat - p3lat
        denom = dx12 * dy34 - dy12 * dx34
        if abs(denom) < 1e-12:
            return None
        dx13 = p3lon - p1lon;  dy13 = p3lat - p1lat
        t = (dx13 * dy34 - dy13 * dx34) / denom
        u = (dx13 * dy12 - dy13 * dx12) / denom
        eps = 1e-9
        if eps < t < 1 - eps and eps < u < 1 - eps:
            return t, u, p1lat + t * dy12, p1lon + t * dx12
        return None

    def _check_new_intersections(self, new_edge: "MapEdge") -> bool:
        """Ищет все рёбра, пересекающиеся с new_edge, создаёт DynamicPoint-ы.
        Возвращает True если хоть одно пересечение найдено."""
        p1 = self._get_point_by_id(new_edge.point_a_id)
        p2 = self._get_point_by_id(new_edge.point_b_id)
        if not p1 or not p2:
            return False
        found = False
        for e in self.map_edges:
            if e is new_edge:
                continue
            # Рёбра с общей вершиной не пересекаются
            if {e.point_a_id, e.point_b_id} & {new_edge.point_a_id, new_edge.point_b_id}:
                continue
            p3 = self._get_point_by_id(e.point_a_id)
            p4 = self._get_point_by_id(e.point_b_id)
            if not p3 or not p4:
                continue
            r = self._segment_intersection(
                p1.lat, p1.lon, p2.lat, p2.lon,
                p3.lat, p3.lon, p4.lat, p4.lon,
            )
            if r is None:
                continue
            key = frozenset([new_edge.id, e.id])
            if any(frozenset([dp.edge_a_id, dp.edge_b_id]) == key
                   for dp in self.dynamic_points):
                continue
            self.dynamic_points.append(
                DynamicPoint(edge_a_id=new_edge.id, edge_b_id=e.id)
            )
            found = True
        return found
    def _remove_stale_dynamic_points(self):
        """Удаляет DynamicPoint-ы, чьи рёбра уже не существуют или не пересекаются."""
        edge_ids = {e.id for e in self.map_edges}
        self.dynamic_points = [
            dp for dp in self.dynamic_points
            if dp.edge_a_id in edge_ids
            and dp.edge_b_id in edge_ids
            and self._compute_dynamic_pos(dp) is not None
        ]

    def _compute_dynamic_pos(self,
                             dp: "DynamicPoint",
                             ) -> "Optional[Tuple[float, float]]":  # type: ignore
        """Вычисляет текущую lat/lon динамической точки пересечения."""
        ea = next((e for e in self.map_edges if e.id == dp.edge_a_id), None)
        eb = next((e for e in self.map_edges if e.id == dp.edge_b_id), None)
        if not ea or not eb:
            return None
        p1 = self._get_point_by_id(ea.point_a_id)
        p2 = self._get_point_by_id(ea.point_b_id)
        p3 = self._get_point_by_id(eb.point_a_id)
        p4 = self._get_point_by_id(eb.point_b_id)
        if not (p1 and p2 and p3 and p4):
            return None
        r = self._segment_intersection(
            p1.lat, p1.lon, p2.lat, p2.lon,
            p3.lat, p3.lon, p4.lat, p4.lon,
        )
        return (r[2], r[3]) if r else None

    def _find_dynamic_point_at(self, sx: float, sy: float,
                               radius: int = 14) -> "Optional[DynamicPoint]":  # type: ignore
        """Возвращает DynamicPoint в радиусе radius px от (sx, sy), или None."""
        for dp in self.dynamic_points:
            pos = self._compute_dynamic_pos(dp)
            if pos is None:
                continue
            dsx, dsy = self._lat_lon_to_screen(pos[0], pos[1])
            if math.hypot(sx - dsx, sy - dsy) <= radius:
                return dp
        return None
    def _split_edge_by_point(self, edge: "MapEdge",
                             new_pt: MapPoint) -> "Tuple[MapEdge, MapEdge]":  # type: ignore
        """Разделяет ребро edge точкой new_pt → два полуребра.
        Обновляет все полигоны, содержащие edge как последовательную пару вершин."""
        a_id, b_id = edge.point_a_id, edge.point_b_id
        if edge in self.map_edges:
            self.map_edges.remove(edge)
        e1 = MapEdge(point_a_id=a_id,      point_b_id=new_pt.id,
                     color=edge.color,     width=edge.width,
                     line_type=edge.line_type, opacity=edge.opacity)
        e2 = MapEdge(point_a_id=new_pt.id, point_b_id=b_id,
                     color=edge.color,     width=edge.width,
                     line_type=edge.line_type, opacity=edge.opacity)
        self.map_edges.extend([e1, e2])
        self.edge_added.emit(e1)
        self.edge_added.emit(e2)
        # Обновляем полигоны, содержащие последовательную пару (a_id, b_id) или (b_id, a_id)
        for poly in self.map_polygons:
            ids = poly.point_ids
            n   = len(ids)
            for i in range(n):
                nxt = (i + 1) % n
                if ids[i] == a_id and ids[nxt] == b_id:
                    poly.point_ids = ids[:i + 1] + [new_pt.id] + ids[i + 1:]
                    break
                elif ids[i] == b_id and ids[nxt] == a_id:
                    poly.point_ids = ids[:i + 1] + [new_pt.id] + ids[i + 1:]
                    break
        return e1, e2
    def _confirm_dynamic_point(self, dp: "DynamicPoint"):
        """Стрл+ЛКМ: подтверждает динамическую точку → реальная MapPoint,
        разбивает оба ребра, разделяет затронутые полигоны."""
        pos = self._compute_dynamic_pos(dp)
        if pos is None:
            return
        ea = next((e for e in self.map_edges if e.id == dp.edge_a_id), None)
        eb = next((e for e in self.map_edges if e.id == dp.edge_b_id), None)
        if not ea or not eb:
            return
        self._push_undo()
        lat, lon = pos
        new_pt = MapPoint(
            lat       = lat,   lon       = lon,
            label     = "",    color     = "#6c7086",
            icon_type = "circle", size   = 0.8,
            is_intersection        = True,
            src_edge_a_endpoints   = (ea.point_a_id, ea.point_b_id),
            src_edge_b_endpoints   = (eb.point_a_id, eb.point_b_id),
        )
        self.map_points.append(new_pt)
        self.point_added.emit(new_pt)
        if dp in self.dynamic_points:
            self.dynamic_points.remove(dp)
        # Делим оба ребра
        self._split_edge_by_point(ea, new_pt)
        self._split_edge_by_point(eb, new_pt)
        # Разделяем полигоны, содержащие new_pt и несмежную соединённую вершину
        for poly in list(self.map_polygons):
            if new_pt.id not in poly.point_ids:
                continue
            ids    = poly.point_ids
            my_idx = ids.index(new_pt.id)
            n      = len(ids)
            for vid in ids:
                if vid == new_pt.id:
                    continue
                vid_idx = ids.index(vid)
                if abs(my_idx - vid_idx) in (0, 1, n - 1):
                    continue  # смежная
                if any({e.point_a_id, e.point_b_id} == {new_pt.id, vid}
                       for e in self.map_edges):
                    self._split_polygon_only(poly, new_pt.id, vid)
                    break
        self.update()
    def _revert_intersection_point(self, pt: MapPoint):
        """Возвращает подтверждённую точку пересечения в режим динамической."""
        self._push_undo()
        a1_id, a2_id = pt.src_edge_a_endpoints
        b1_id, b2_id = pt.src_edge_b_endpoints
        # Собираем полурёбра, связанные с pt
        half_edges = [e for e in self.map_edges
                      if e.point_a_id == pt.id or e.point_b_id == pt.id]
        # Подбираем цвет/стиль из полурёбер (берём первое из каждой пары)
        kw_a = dict(color=self.default_edge_color, width=self.default_edge_width,
                    line_type=self.default_edge_type, opacity=self.default_edge_opacity)
        kw_b = dict(**kw_a)
        for e in half_edges:
            other = e.point_b_id if e.point_a_id == pt.id else e.point_a_id
            if other in (a1_id, a2_id):
                kw_a = dict(color=e.color, width=e.width,
                            line_type=e.line_type, opacity=e.opacity)
            elif other in (b1_id, b2_id):
                kw_b = dict(color=e.color, width=e.width,
                            line_type=e.line_type, opacity=e.opacity)
        for e in half_edges:
            if e in self.map_edges:
                self.map_edges.remove(e)
        # Восстанавливаем исходные рёбра
        ea = MapEdge(point_a_id=a1_id, point_b_id=a2_id, **kw_a)
        eb = MapEdge(point_a_id=b1_id, point_b_id=b2_id, **kw_b)
        self.map_edges.extend([ea, eb])
        # Убираем pt из полигонов
        for poly in self.map_polygons:
            if pt.id in poly.point_ids:
                poly.point_ids = [vid for vid in poly.point_ids if vid != pt.id]
        self.map_polygons = [p for p in self.map_polygons if len(p.point_ids) >= 3]
        # Удаляем саму точку
        if pt in self.map_points:
            self.map_points.remove(pt)
        self.selected_point_ids.discard(pt.id)
        # Добавляем DynamicPoint для восстановленного пересечения
        self.dynamic_points.append(DynamicPoint(edge_a_id=ea.id, edge_b_id=eb.id))
        self.update()

    def _draw_dynamic_points(self, painter: QPainter):
        """Рисует динамические точки пересечений (серые, пунктирные круги)."""
        if not self.show_dynamic_points:
            return
        self._remove_stale_dynamic_points()
        W, H = self.width(), self.height()
        fnt  = QFont("Segoe UI", 7)
        painter.setFont(fnt)
        for dp in self.dynamic_points:
            pos = self._compute_dynamic_pos(dp)
            if pos is None:
                continue
            sx, sy = self._lat_lon_to_screen(pos[0], pos[1])
            if sx < -20 or sy < -20 or sx > W + 20 or sy > H + 20:
                continue
            r = 8
            painter.setOpacity(0.85)
            painter.setPen(QPen(QColor("#6c7086"), 1.5, Qt.PenStyle.DashLine))
            painter.setBrush(QBrush(QColor("#45475a")))
            painter.drawEllipse(sx - r, sy - r, 2 * r, 2 * r)
            painter.setPen(QPen(QColor("#9399b2"), 1.0))
            painter.drawLine(sx - 4, sy, sx + 4, sy)
            painter.drawLine(sx, sy - 4, sx, sy + 4)
    # ── Разрезание полигона ───────────────────────────────────────

    def _split_polygon_only(self, poly: "MapPolygon",
                            pt1_id: str, pt2_id: str) -> bool:
        """Внутренний сплит: только меняет полигоны, не добавляет ребро, без undo."""
        ids = poly.point_ids
        n   = len(ids)
        if n < 4 or pt1_id not in ids or pt2_id not in ids:
            return False
        i1, i2 = ids.index(pt1_id), ids.index(pt2_id)
        if i1 > i2:
            i1, i2 = i2, i1
        sub1 = ids[i1:i2 + 1]
        sub2 = ids[i2:] + ids[:i1 + 1]
        if len(sub1) < 3 or len(sub2) < 3:
            return False
        kw = dict(
            fill_color   = poly.fill_color,   fill_opacity  = poly.fill_opacity,
            border_color = poly.border_color, border_width  = poly.border_width,
            border_type  = poly.border_type,
        )
        self.map_polygons.remove(poly)
        self.map_polygons.append(MapPolygon(point_ids=sub1, **kw))
        self.map_polygons.append(MapPolygon(point_ids=sub2, **kw))
        return True

    def split_polygon(self, poly: "MapPolygon", pt1_id: str, pt2_id: str) -> bool:
        """Разрезает полигон по двум точкам: добавляет ребро + разбивает на 2 полигона."""
        ids = poly.point_ids
        if (len(ids) < 4 or pt1_id not in ids or pt2_id not in ids
                or abs(ids.index(pt1_id) - ids.index(pt2_id)) in (0, 1, len(ids) - 1)):
            return False   # смежные или идентичные — не разрезаем
        # Проверяем что sub1 и sub2 >= 3
        i1, i2 = ids.index(pt1_id), ids.index(pt2_id)
        if i1 > i2:
            i1, i2 = i2, i1
        if len(ids[i1:i2 + 1]) < 3 or len(ids[i2:] + ids[:i1 + 1]) < 3:
            return False
        self._push_undo()
        # Добавляем ребро-разрез напрямую (без _create_edge, чтобы не зациклиться)
        if not any({e.point_a_id, e.point_b_id} == {pt1_id, pt2_id}
                   for e in self.map_edges):
            edge = MapEdge(
                point_a_id = pt1_id, point_b_id = pt2_id,
                color      = self.default_edge_color,
                width      = self.default_edge_width,
                line_type  = self.default_edge_type,
                opacity    = self.default_edge_opacity,
            )
            self.map_edges.append(edge)
            self.edge_added.emit(edge)
        self._split_polygon_only(poly, pt1_id, pt2_id)
        self.update()
        return True

    def _add_map_point_at(self, lat: float, lon: float):
        """Создаёт новую точку с текущими настройками по умолчанию."""
        self._push_undo()
        pt = MapPoint(
            lat       = lat,
            lon       = lon,
            icon_type = self.default_icon_type,
            color     = self.default_color,
            size      = self.default_size,
        )
        self.map_points.append(pt)
        self.selected_point_ids = {pt.id}
        self.point_added.emit(pt)
        self.selection_changed.emit([pt])
        # Авто-соединение: к последней поставленной точке
        if self.auto_connect:
            # Проверяем: новая точка внутри полигона?
            host_poly = next(
                (p for p in self.map_polygons
                 if self._point_in_polygon_latlon(pt.lat, pt.lon, p)),
                None,
            )
            if host_poly is not None:
                # Соединяем с ВСЕМИ вершинами полигона (без авто-пересечений/полигонов)
                for vid in list(host_poly.point_ids):
                    self._create_edge_raw(pt.id, vid)
                # Разбиваем полигон на секторы
                self._subdivide_by_interior(host_poly, pt.id,
                                            list(host_poly.point_ids))
            else:
                if self._last_placed_point_id is not None:
                    if self._get_point_by_id(self._last_placed_point_id) is not None:
                        self._create_edge(self._last_placed_point_id, pt.id)
                else:
                    self._chain_start_id = pt.id   # первая точка новой цепи
        self._last_placed_point_id = pt.id
        self._sync_dynamic_points()
        self.update()

    def _delete_map_point(self, pt: MapPoint):
        """Удаляет точку, все связанные рёбра и полигоны."""
        self._push_undo()
        if self._last_placed_point_id == pt.id:
            self._last_placed_point_id = None
        # Удаляем рёбра, затрагивающие эту точку
        edges_to_remove = [
            e for e in self.map_edges
            if e.point_a_id == pt.id or e.point_b_id == pt.id
        ]
        for e in edges_to_remove:
            self.map_edges.remove(e)
            self.edge_deleted.emit(e)
        # Удаляем полигоны, содержащие эту точку
        self.map_polygons = [
            poly for poly in self.map_polygons if pt.id not in poly.point_ids
        ]
        # Удаляем саму точку
        if pt in self.map_points:
            self.map_points.remove(pt)
        self.selected_point_ids.discard(pt.id)
        if self._hovered_point_id == pt.id:
            self._hovered_point_id = None
        if self._pivot_point_id == pt.id:
            self._pivot_point_id = None
        if self._connect_source_id == pt.id:
            self._connect_source_id = None
        self.selection_changed.emit(self._get_selected_points())
        self.point_deleted.emit(pt)
        self.update()

    def _show_point_ctx_menu(self, pt: MapPoint, global_pos):
        """Показывает контекстное меню для точки."""
        menu = QMenu(self)
        menu.setStyleSheet(
            "QMenu { background:#1e1e2e; color:#cdd6f4; border:1px solid #45475a; "
            "padding:2px; } "
            "QMenu::item { padding:4px 16px; } "
            "QMenu::item:selected { background:#313244; } "
            "QMenu::separator { background:#45475a; height:1px; margin:3px 8px; }"
        )
        is_selected = pt.id in self.selected_point_ids
        is_pivot    = pt.id == self._pivot_point_id
        n_sel       = len(self.selected_point_ids)

        if is_selected and n_sel > 1:
            act_sel = menu.addAction("Выделить только эту точку")
        else:
            act_sel = menu.addAction("Выделить точку")

        pivot_label = "Снять точку опоры" if is_pivot else "Сделать точкой опоры"
        act_pivot = menu.addAction(pivot_label)
        menu.addSeparator()
        act_copy = menu.addAction("Копировать координаты")
        menu.addSeparator()

        act_del_sel = None
        if n_sel > 1 and is_selected:
            act_del_sel = menu.addAction(f"Удалить выбранные ({n_sel})")
        act_del = menu.addAction("Удалить точку")

        # Если точка — подтверждённое пересечение, добавляем специальные опции
        act_revert = None
        if pt.is_intersection:
            menu.addSeparator()
            act_revert = menu.addAction("↩ Вернуть как динамическую точку")

        chosen = menu.exec(global_pos)
        if act_revert is not None and chosen == act_revert:
            self._revert_intersection_point(pt)
        elif chosen == act_del:
            self._delete_map_point(pt)
        elif act_del_sel is not None and chosen == act_del_sel:
            for p in list(self._get_selected_points()):
                self._delete_map_point(p)
        elif chosen == act_copy:
            QApplication.clipboard().setText(f"{pt.lat:.6f}, {pt.lon:.6f}")
        elif chosen == act_sel:
            self.selected_point_ids = {pt.id}
            self.selection_changed.emit([pt])
            self.update()
        elif chosen == act_pivot:
            self._pivot_point_id = None if is_pivot else pt.id
            self.update()


# ══════════════════════════════════════════════════════════════
#  Вкладка «Карта»
# ══════════════════════════════════════════════════════════════

# Классификатор слоёв: (подстрока в имени файла, отображаемое имя, прозрачность, порядок)
# Всё что не совпало → базовый слой (порядок 1)
_LAYER_PATTERNS: List[Tuple[str, str, float, int]] = [
    ("fgis",        "Кадастр ФГИС",  0.7, 3),
    ("miass_alpha", "OSM",           1.0, 2),
]


class MapTab(QWidget):
    """Вкладка карты с выбором координат и вставкой в таблицу данных."""

    def __init__(self, cfg: Config, data_tab):
        super().__init__()
        self.cfg      = cfg
        self.data_tab = data_tab

        self._providers: List[TileProvider] = []
        self._insert_sequential = True
        self._next_row_idx      = 0
        self._last_lat: Optional[float] = None
        self._last_lon: Optional[float] = None

        self.insert_g  = None   # сохраняется в _setup_ui для show/hide
        self.measure_g = None   # группа измерения (view mode)

        self._setup_ui()
        self._load_databases()

    # ── Построение UI ────────────────────────────────────────────

    def _setup_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Горизонтальный сплиттер: левый (карта) | правый (список объектов)
        self._main_splitter = QSplitter(Qt.Orientation.Horizontal)
        self._main_splitter.setChildrenCollapsible(True)
        self._main_splitter.setHandleWidth(4)

        # Левая часть: канвас + нижняя панель
        _left_w = QWidget()
        _left_l = QVBoxLayout(_left_w)
        _left_l.setContentsMargins(0, 0, 0, 0)
        _left_l.setSpacing(0)
        self._left_l = _left_l   # сохраняем для дальнейших addWidget

        # ── Карта ────────────────────────────────────────────────
        self.canvas = MapCanvas(self)
        self.canvas.coord_clicked.connect(self._on_click)
        self.canvas.coord_hovered.connect(self._on_hover)
        self.canvas._update_timer.timeout.connect(self._sync_zoom_label)
        self.canvas.temp_pts_changed.connect(self._update_distance_label)
        self.canvas.point_added.connect(self._on_point_changed)
        self.canvas.point_moved.connect(self._on_point_changed)
        self.canvas.point_deleted.connect(self._on_point_deleted)
        self.canvas.selection_changed.connect(self._on_selection_changed)
        self.canvas.edge_added.connect(lambda _: self._update_geo_labels())
        self.canvas.edge_deleted.connect(lambda _: self._update_geo_labels())
        self.canvas.labels_toggled.connect(self._on_labels_toggled)
        self.canvas.point_added.connect(lambda _: self._refresh_obj_list())
        self.canvas.point_deleted.connect(lambda _: self._refresh_obj_list())
        self.canvas.point_moved.connect(lambda _: self._refresh_obj_list())
        self.canvas.edge_added.connect(lambda _: self._refresh_obj_list())
        self.canvas.edge_deleted.connect(lambda _: self._refresh_obj_list())
        self._left_l.addWidget(self.canvas, 1)

        # ── Нижняя панель управления ─────────────────────────────
        panel = QWidget()
        panel.setStyleSheet(
            "QWidget { background-color: #181825; }"
            "QWidget#map_panel { border-top: 1px solid #313244; }"
        )
        panel.setObjectName("map_panel")

        pl = QHBoxLayout(panel)
        pl.setContentsMargins(10, 6, 10, 6)
        pl.setSpacing(10)

        _gs = (
            "QGroupBox { font-size:9pt; color:#a6adc8; border:1px solid #45475a; "
            "border-radius:4px; margin-top:6px; padding-top:8px; } "
            "QGroupBox::title { subcontrol-origin:margin; left:8px; padding:0 4px; }"
        )

        # ─── Поиск координат ─────────────────────────────────────
        search_g = QGroupBox("Поиск координат")
        search_g.setStyleSheet(_gs)
        sg_l = QVBoxLayout(search_g)
        sg_l.setContentsMargins(6, 10, 6, 6)
        sg_l.setSpacing(4)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("55.123456, 60.654321")
        self.search_edit.setMinimumWidth(160)
        self.search_edit.returnPressed.connect(self._on_search)
        sg_l.addWidget(self.search_edit)

        self.btn_search = QPushButton("Перейти")
        self.btn_search.clicked.connect(self._on_search)
        sg_l.addWidget(self.btn_search)

        self.chk_search_add_point = QCheckBox("С точкой")
        self.chk_search_add_point.setChecked(False)
        sg_l.addWidget(self.chk_search_add_point)

        pl.addWidget(search_g)

        # ─── Слои ────────────────────────────────────────────────
        self.layers_g  = QGroupBox("Слои")
        self.layers_g.setStyleSheet(_gs)
        self.layers_vl = QVBoxLayout(self.layers_g)
        self.layers_vl.setContentsMargins(6, 10, 6, 4)
        self.layers_vl.setSpacing(3)
        pl.addWidget(self.layers_g, 1)   # растягиваем

        # ─── Навигация ────────────────────────────────────────────
        nav_g = QGroupBox("Навигация")
        nav_g.setStyleSheet(_gs)
        nav_l = QVBoxLayout(nav_g)
        nav_l.setContentsMargins(6, 10, 6, 6)
        nav_l.setSpacing(4)

        # Зум
        zoom_w = QWidget()
        zoom_l = QHBoxLayout(zoom_w)
        zoom_l.setContentsMargins(0, 0, 0, 0)
        zoom_l.setSpacing(4)

        self.btn_zoom_out = QPushButton("−")
        self.btn_zoom_out.setFixedSize(28, 28)
        self.zoom_lbl = QLabel("12")
        self.zoom_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.zoom_lbl.setFixedWidth(30)
        self.zoom_lbl.setStyleSheet("color:#cdd6f4; font-family:Consolas;")
        self.btn_zoom_in = QPushButton("+")
        self.btn_zoom_in.setFixedSize(28, 28)

        self.btn_zoom_out.clicked.connect(lambda: self.canvas.set_zoom(self.canvas.zoom - 1))
        self.btn_zoom_in.clicked.connect(lambda: self.canvas.set_zoom(self.canvas.zoom + 1))

        zoom_l.addWidget(self.btn_zoom_out)
        zoom_l.addWidget(self.zoom_lbl)
        zoom_l.addWidget(self.btn_zoom_in)
        zoom_l.addStretch()
        nav_l.addWidget(zoom_w)

        # Переключатель режима
        mode_lbl = QLabel("Режим:")
        mode_lbl.setStyleSheet("color:#a6adc8; font-size:9pt;")
        nav_l.addWidget(mode_lbl)

        mode_w = QWidget()
        mode_l = QHBoxLayout(mode_w)
        mode_l.setContentsMargins(0, 0, 0, 0)
        mode_l.setSpacing(6)

        self.radio_edit = QRadioButton("Редактирование")
        self.radio_view = QRadioButton("Обзор / Измерение")
        self.radio_pts  = QRadioButton("Точки")
        self.radio_edit.setChecked(True)

        mode_btn_grp = QButtonGroup(self)
        mode_btn_grp.addButton(self.radio_edit)
        mode_btn_grp.addButton(self.radio_view)
        mode_btn_grp.addButton(self.radio_pts)
        self.radio_edit.toggled.connect(self._on_mode_changed)
        self.radio_view.toggled.connect(self._on_mode_changed)
        self.radio_pts.toggled.connect(self._on_mode_changed)

        mode_l.addWidget(self.radio_edit)
        mode_l.addWidget(self.radio_view)
        mode_l.addWidget(self.radio_pts)
        nav_l.addWidget(mode_w)

        self.btn_center = QPushButton("По центру покрытия")
        self.btn_center.clicked.connect(self._center_on_coverage)
        nav_l.addWidget(self.btn_center)
        nav_l.addStretch()

        pl.addWidget(nav_g)

        # ─── Вставка координат в таблицу ─────────────────────────
        self.insert_g = QGroupBox("Вставка координат в таблицу")
        self.insert_g.setStyleSheet(_gs)
        insert_l = QVBoxLayout(self.insert_g)
        insert_l.setContentsMargins(6, 10, 6, 6)
        insert_l.setSpacing(4)

        ins_mode_w = QWidget()
        ins_mode_l = QHBoxLayout(ins_mode_w)
        ins_mode_l.setContentsMargins(0, 0, 0, 0)
        ins_mode_l.setSpacing(8)

        self.radio_seq = QRadioButton("Посл-но")
        self.radio_sel = QRadioButton("В выбранную")
        self.radio_seq.setChecked(True)
        btn_grp = QButtonGroup(self)
        btn_grp.addButton(self.radio_seq)
        btn_grp.addButton(self.radio_sel)
        self.radio_seq.toggled.connect(self._on_insert_mode_changed)

        ins_mode_l.addWidget(self.radio_seq)
        ins_mode_l.addWidget(self.radio_sel)
        insert_l.addWidget(ins_mode_w)

        self.lbl_insert_status = QLabel("Следующая строка: 1")
        self.lbl_insert_status.setStyleSheet("color:#a6adc8; font-size:9pt;")
        insert_l.addWidget(self.lbl_insert_status)

        btns_w = QWidget()
        btns_l = QHBoxLayout(btns_w)
        btns_l.setContentsMargins(0, 0, 0, 0)
        btns_l.setSpacing(4)
        self.btn_reset_seq      = QPushButton("Сброс")
        self.btn_update_markers = QPushButton("Обновить маркеры")
        self.btn_reset_seq.clicked.connect(self._reset_sequential)
        self.btn_update_markers.clicked.connect(self._refresh_markers)
        btns_l.addWidget(self.btn_reset_seq)
        btns_l.addWidget(self.btn_update_markers)
        insert_l.addWidget(btns_w)
        insert_l.addStretch()

        pl.addWidget(self.insert_g)

        # ─── Группа измерения (только в режиме обзора) ───────────
        self.measure_g = QGroupBox("Измерение расстояний")
        self.measure_g.setStyleSheet(_gs)
        meas_l = QVBoxLayout(self.measure_g)
        meas_l.setContentsMargins(6, 10, 6, 6)
        meas_l.setSpacing(4)

        self.lbl_distance = QLabel("Точек: 0")
        self.lbl_distance.setStyleSheet(
            "color:#f9e2af; font-family:Consolas; font-size:9pt;"
        )
        meas_l.addWidget(self.lbl_distance)

        lbl_hint = QLabel("ЛКМ — добавить точку\nПКМ — удалить последнюю")
        lbl_hint.setStyleSheet("color:#6c7086; font-size:8pt;")
        meas_l.addWidget(lbl_hint)

        self.btn_clear_measure = QPushButton("Очистить точки")
        self.btn_clear_measure.clicked.connect(self._clear_temp_points)
        meas_l.addWidget(self.btn_clear_measure)
        meas_l.addStretch()

        self.measure_g.setVisible(False)  # скрыта по умолчанию
        pl.addWidget(self.measure_g)

        # ─── Панель свойств точек ─────────────────────────────────
        self.points_g = QGroupBox("Точки")
        self.points_g.setStyleSheet(_gs)
        pts_l = QVBoxLayout(self.points_g)
        pts_l.setContentsMargins(6, 10, 6, 6)
        pts_l.setSpacing(4)

        # Строка: имя точки
        lbl_name = QLabel("Имя:")
        lbl_name.setStyleSheet("color:#a6adc8; font-size:9pt;")
        self.pt_name_edit = QLineEdit()
        self.pt_name_edit.setPlaceholderText("название точки")
        self.pt_name_edit.setMaximumWidth(130)
        self.pt_name_edit.editingFinished.connect(self._apply_point_name)
        name_row = QWidget()
        name_rl  = QHBoxLayout(name_row)
        name_rl.setContentsMargins(0, 0, 0, 0)
        name_rl.setSpacing(4)
        name_rl.addWidget(lbl_name)
        name_rl.addWidget(self.pt_name_edit)
        pts_l.addWidget(name_row)

        # Строка: иконка + цвет
        icon_row = QWidget()
        icon_rl  = QHBoxLayout(icon_row)
        icon_rl.setContentsMargins(0, 0, 0, 0)
        icon_rl.setSpacing(4)

        lbl_icon = QLabel("Иконка:")
        lbl_icon.setStyleSheet("color:#a6adc8; font-size:9pt;")
        self.pt_icon_combo = QComboBox()
        self.pt_icon_combo.setMaximumWidth(105)
        for ru_name in PRESET_ICONS_RU:
            self.pt_icon_combo.addItem(ru_name)
        self.pt_icon_combo.addItem("Свой...")
        self.pt_icon_combo.currentIndexChanged.connect(self._on_icon_changed)

        self.pt_color_btn = QPushButton()
        self.pt_color_btn.setFixedSize(24, 24)
        self.pt_color_btn.setToolTip("Цвет точки")
        self.pt_color_btn.clicked.connect(self._pick_point_color)
        self._set_color_btn(self.canvas.default_color)

        icon_rl.addWidget(lbl_icon)
        icon_rl.addWidget(self.pt_icon_combo)
        icon_rl.addWidget(self.pt_color_btn)
        pts_l.addWidget(icon_row)

        # Строка: размер
        size_row = QWidget()
        size_rl  = QHBoxLayout(size_row)
        size_rl.setContentsMargins(0, 0, 0, 0)
        size_rl.setSpacing(4)

        lbl_size = QLabel("Размер:")
        lbl_size.setStyleSheet("color:#a6adc8; font-size:9pt;")
        self.pt_size_slider = QSlider(Qt.Orientation.Horizontal)
        self.pt_size_slider.setRange(5, 30)   # 0.5 – 3.0 с шагом 0.1
        self.pt_size_slider.setValue(10)       # = 1.0
        self.pt_size_slider.setFixedWidth(80)
        self.pt_size_lbl = QLabel("1.0×")
        self.pt_size_lbl.setStyleSheet("color:#a6adc8; font-size:8pt;")
        self.pt_size_lbl.setFixedWidth(30)
        self.pt_size_slider.valueChanged.connect(self._on_size_changed)

        size_rl.addWidget(lbl_size)
        size_rl.addWidget(self.pt_size_slider)
        size_rl.addWidget(self.pt_size_lbl)
        pts_l.addWidget(size_row)

        # Строка: вращение выделения
        rot_row = QWidget()
        rot_rl  = QHBoxLayout(rot_row)
        rot_rl.setContentsMargins(0, 0, 0, 0)
        rot_rl.setSpacing(2)
        lbl_rot = QLabel("Поворот:")
        lbl_rot.setStyleSheet("color:#a6adc8; font-size:9pt;")
        rot_rl.addWidget(lbl_rot)
        for deg in (-90, -15, -5, -1, +1, +5, +15, +90):
            sign  = "+" if deg > 0 else ""
            b     = QPushButton(f"{sign}{deg}°")
            b.setFixedHeight(22)
            b.setFixedWidth(38)
            b.clicked.connect(lambda _checked, d=deg: self._rotate_selection(d))
            rot_rl.addWidget(b)
        rot_rl.addStretch()
        self.lbl_pivot = QLabel("Опора: —")
        self.lbl_pivot.setStyleSheet("color:#6c7086; font-size:8pt;")
        rot_rl.addWidget(self.lbl_pivot)
        pts_l.addWidget(rot_row)

        # Строка: счётчик + подсказка
        self.lbl_pts_count = QLabel("Точек на карте: 0")
        self.lbl_pts_count.setStyleSheet("color:#a6adc8; font-size:9pt;")
        pts_l.addWidget(self.lbl_pts_count)

        lbl_pts_hint = QLabel(
            "Ctrl+ЛКМ — добавить точку  |  ЛКМ+тащи — переместить\n"
            "ЛКМ — выбрать  |  Shift+ЛКМ — мульти-выбор\n"
            "Ctrl+ПКМ — удалить  |  Alt+ЛКМ — опора поворота\n"
            "R (зажать) — поворот мышью  |  Esc — сброс"
        )
        lbl_pts_hint.setStyleSheet("color:#6c7086; font-size:8pt;")
        pts_l.addWidget(lbl_pts_hint)

        self.points_g.setVisible(False)   # скрыта по умолчанию
        pl.addWidget(self.points_g)

        # ─── Панель рёбер ─────────────────────────────────────────
        self.edges_g = QGroupBox("Рёбра")
        self.edges_g.setStyleSheet(_gs)
        edg_l = QVBoxLayout(self.edges_g)
        edg_l.setContentsMargins(6, 10, 6, 6)
        edg_l.setSpacing(4)

        # Тип + цвет + толщина
        ec_row = QWidget(); ec_rl = QHBoxLayout(ec_row)
        ec_rl.setContentsMargins(0, 0, 0, 0); ec_rl.setSpacing(4)
        lbl_etype = QLabel("Тип:")
        lbl_etype.setStyleSheet("color:#a6adc8; font-size:9pt;")
        self.edge_type_combo = QComboBox()
        self.edge_type_combo.setMaximumWidth(90)
        for ru in EDGE_TYPES_RU:
            self.edge_type_combo.addItem(ru)
        self.edge_type_combo.currentIndexChanged.connect(self._on_edge_type_changed)
        self.edge_color_btn = QPushButton()
        self.edge_color_btn.setFixedSize(22, 22)
        self.edge_color_btn.setToolTip("Цвет ребра")
        self.edge_color_btn.clicked.connect(self._pick_edge_color)
        self._set_edge_color_btn(self.canvas.default_edge_color)
        ec_rl.addWidget(lbl_etype)
        ec_rl.addWidget(self.edge_type_combo)
        ec_rl.addWidget(self.edge_color_btn)
        ec_rl.addStretch()
        edg_l.addWidget(ec_row)

        # Толщина
        ew_row = QWidget(); ew_rl = QHBoxLayout(ew_row)
        ew_rl.setContentsMargins(0, 0, 0, 0); ew_rl.setSpacing(4)
        lbl_ew = QLabel("Толщина:")
        lbl_ew.setStyleSheet("color:#a6adc8; font-size:9pt;")
        self.edge_width_slider = QSlider(Qt.Orientation.Horizontal)
        self.edge_width_slider.setRange(1, 20)
        self.edge_width_slider.setValue(int(self.canvas.default_edge_width * 2))
        self.edge_width_slider.setFixedWidth(70)
        self.edge_width_lbl = QLabel(f"{self.canvas.default_edge_width:.1f}")
        self.edge_width_lbl.setStyleSheet("color:#a6adc8; font-size:8pt;")
        self.edge_width_lbl.setFixedWidth(28)
        self.edge_width_slider.valueChanged.connect(self._on_edge_width_changed)
        ew_rl.addWidget(lbl_ew)
        ew_rl.addWidget(self.edge_width_slider)
        ew_rl.addWidget(self.edge_width_lbl)
        edg_l.addWidget(ew_row)

        # Авто-соединение + кнопки
        eco_row = QWidget(); eco_rl = QHBoxLayout(eco_row)
        eco_rl.setContentsMargins(0, 0, 0, 0); eco_rl.setSpacing(4)
        self.chk_auto_connect = QCheckBox("Авто")
        self.chk_auto_connect.setToolTip("Авто-соединение при добавлении точки")
        self.chk_auto_connect.setChecked(self.canvas.auto_connect)
        self.chk_auto_connect.toggled.connect(lambda v: setattr(self.canvas, "auto_connect", v))
        self.btn_connect      = QPushButton("Соединить")
        self.btn_connect.setFixedHeight(22)
        self.btn_connect.clicked.connect(self._connect_selected_points)
        self.btn_disconnect   = QPushButton("Разъединить")
        self.btn_disconnect.setFixedHeight(22)
        self.btn_disconnect.clicked.connect(self._disconnect_selected_points)
        eco_rl.addWidget(self.chk_auto_connect)
        eco_rl.addWidget(self.btn_connect)
        eco_rl.addWidget(self.btn_disconnect)
        edg_l.addWidget(eco_row)

        # Ручной режим соединения
        self.btn_connect_mode = QPushButton("Ручное соединение")
        self.btn_connect_mode.setCheckable(True)
        self.btn_connect_mode.setFixedHeight(22)
        self.btn_connect_mode.toggled.connect(self._on_connect_mode_toggled)
        edg_l.addWidget(self.btn_connect_mode)

        self.edges_g.setVisible(False)
        pl.addWidget(self.edges_g)

        # ─── Панель полигонов ─────────────────────────────────────
        self.polygon_g = QGroupBox("Полигоны")
        self.polygon_g.setStyleSheet(_gs)
        poly_l = QVBoxLayout(self.polygon_g)
        poly_l.setContentsMargins(6, 10, 6, 6)
        poly_l.setSpacing(4)

        # Цвет заливки + прозрачность
        pf_row = QWidget(); pf_rl = QHBoxLayout(pf_row)
        pf_rl.setContentsMargins(0, 0, 0, 0); pf_rl.setSpacing(4)
        lbl_pf = QLabel("Заливка:")
        lbl_pf.setStyleSheet("color:#a6adc8; font-size:9pt;")
        self.poly_fill_btn = QPushButton()
        self.poly_fill_btn.setFixedSize(22, 22)
        self.poly_fill_btn.setToolTip("Цвет заливки полигона")
        self.poly_fill_btn.clicked.connect(self._pick_poly_fill_color)
        self._set_poly_fill_btn("#89b4fa")
        self.poly_fill_slider = QSlider(Qt.Orientation.Horizontal)
        self.poly_fill_slider.setRange(0, 100)
        self.poly_fill_slider.setValue(25)
        self.poly_fill_slider.setFixedWidth(60)
        self.poly_fill_opacity_lbl = QLabel("25%")
        self.poly_fill_opacity_lbl.setStyleSheet("color:#a6adc8; font-size:8pt;")
        self.poly_fill_opacity_lbl.setFixedWidth(30)
        self.poly_fill_slider.valueChanged.connect(self._on_poly_fill_opacity_changed)
        pf_rl.addWidget(lbl_pf)
        pf_rl.addWidget(self.poly_fill_btn)
        pf_rl.addWidget(self.poly_fill_slider)
        pf_rl.addWidget(self.poly_fill_opacity_lbl)
        poly_l.addWidget(pf_row)

        # Цвет рамки + тип
        pb_row = QWidget(); pb_rl = QHBoxLayout(pb_row)
        pb_rl.setContentsMargins(0, 0, 0, 0); pb_rl.setSpacing(4)
        lbl_pb = QLabel("Рамка:")
        lbl_pb.setStyleSheet("color:#a6adc8; font-size:9pt;")
        self.poly_border_btn = QPushButton()
        self.poly_border_btn.setFixedSize(22, 22)
        self.poly_border_btn.setToolTip("Цвет рамки полигона")
        self.poly_border_btn.clicked.connect(self._pick_poly_border_color)
        self._set_poly_border_btn("#89b4fa")
        self.poly_border_type_combo = QComboBox()
        self.poly_border_type_combo.setMaximumWidth(90)
        for ru in EDGE_TYPES_RU:
            self.poly_border_type_combo.addItem(ru)
        self.poly_border_type_combo.currentIndexChanged.connect(self._on_poly_border_type_changed)
        pb_rl.addWidget(lbl_pb)
        pb_rl.addWidget(self.poly_border_btn)
        pb_rl.addWidget(self.poly_border_type_combo)
        poly_l.addWidget(pb_row)

        # Авто-полигон + создать + метки
        popt_row = QWidget(); popt_rl = QHBoxLayout(popt_row)
        popt_rl.setContentsMargins(0, 0, 0, 0); popt_rl.setSpacing(4)
        self.chk_auto_polygon = QCheckBox("Авто-полигон")
        self.chk_auto_polygon.setChecked(self.canvas.auto_polygon)
        self.chk_auto_polygon.toggled.connect(lambda v: setattr(self.canvas, "auto_polygon", v))
        self.btn_labels_filter = QToolButton()
        self.btn_labels_filter.setText("Метки ▾")
        self.btn_labels_filter.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.btn_labels_filter.setToolTip("Фильтр отображения меток на карте")
        self._build_labels_filter_menu()
        self.btn_create_poly  = QPushButton("Создать")
        self.btn_create_poly.setFixedHeight(22)
        self.btn_create_poly.setToolTip("Создать полигон из выбранных точек")
        self.btn_create_poly.clicked.connect(self._create_polygon_from_selected)
        self.btn_split_poly = QPushButton("Разделить")
        self.btn_split_poly.setFixedHeight(22)
        self.btn_split_poly.setToolTip("Разрезать полигон по 2 выбранным точкам")
        self.btn_split_poly.clicked.connect(self._split_selected_polygon)
        popt_rl.addWidget(self.chk_auto_polygon)
        popt_rl.addWidget(self.btn_labels_filter)
        popt_rl.addWidget(self.btn_create_poly)
        popt_rl.addWidget(self.btn_split_poly)
        poly_l.addWidget(popt_row)

        self.poly_import_edit = QTextEdit()
        self.poly_import_edit.setPlaceholderText(
            "Импорт полигона: каждая строка — точка (широта, долгота).\n"
            "Разделители: запятая, пробел, таб, точка после числа, суффиксы N/E."
        )
        self.poly_import_edit.setFixedHeight(84)
        poly_l.addWidget(self.poly_import_edit)

        self.chk_auto_name_points = QCheckBox("Авто имя точек")
        self.chk_auto_name_points.setToolTip("Присвоить точкам имена 1..N при создании/импорте")
        self.chk_auto_name_points.setChecked(True)
        poly_l.addWidget(self.chk_auto_name_points)

        pimp_row = QWidget(); pimp_rl = QHBoxLayout(pimp_row)
        pimp_rl.setContentsMargins(0, 0, 0, 0); pimp_rl.setSpacing(4)
        self.btn_import_poly_text = QPushButton("Импорт текста")
        self.btn_import_poly_text.setToolTip("Создать полигон из координат в поле выше")
        self.btn_import_poly_text.clicked.connect(self._import_polygon_from_text)
        self.btn_import_poly_file = QPushButton("Импорт файла")
        self.btn_import_poly_file.clicked.connect(self._import_polygon_from_file)
        self.btn_make_import_examples = QPushButton("Примеры")
        self.btn_make_import_examples.setToolTip("Создать примеры файлов импорта (txt/json/xml/xlsx)")
        self.btn_make_import_examples.clicked.connect(self._create_polygon_import_examples)
        pimp_rl.addWidget(self.btn_import_poly_text)
        pimp_rl.addWidget(self.btn_import_poly_file)
        pimp_rl.addWidget(self.btn_make_import_examples)
        poly_l.addWidget(pimp_row)

        self.lbl_poly_import_status = QLabel("Импорт: вставьте минимум 3 точки")
        self.lbl_poly_import_status.setStyleSheet("color:#6c7086; font-size:8pt;")
        poly_l.addWidget(self.lbl_poly_import_status)

        # Динамические точки/разделение при пересечениях
        dyn_row = QWidget(); dyn_rl = QHBoxLayout(dyn_row)
        dyn_rl.setContentsMargins(0, 0, 0, 0); dyn_rl.setSpacing(4)
        self.chk_show_dynamic_points = QCheckBox("Дин. точки")
        self.chk_show_dynamic_points.setToolTip("Показывать динамические точки пересечений")
        self.chk_show_dynamic_points.setChecked(self.canvas.show_dynamic_points)
        self.chk_show_dynamic_points.toggled.connect(self._on_show_dynamic_points_toggled)
        self.chk_auto_dynamic_points = QCheckBox("Создавать")
        self.chk_auto_dynamic_points.setToolTip("Создавать динамические точки при любых пересечениях рёбер")
        self.chk_auto_dynamic_points.setChecked(self.canvas.auto_dynamic_points)
        self.chk_auto_dynamic_points.toggled.connect(self._on_auto_dynamic_points_toggled)
        dyn_rl.addWidget(self.chk_show_dynamic_points)
        dyn_rl.addWidget(self.chk_auto_dynamic_points)
        dyn_rl.addStretch()
        poly_l.addWidget(dyn_row)

        self.polygon_g.setVisible(False)
        pl.addWidget(self.polygon_g)

        # ─── Геометрия (суммарная статистика) ────────────────────
        self.geo_g = QGroupBox("Геометрия")
        self.geo_g.setStyleSheet(_gs)
        geo_l = QVBoxLayout(self.geo_g)
        geo_l.setContentsMargins(6, 10, 6, 6)
        geo_l.setSpacing(4)

        self.lbl_total_length = QLabel("Рёбра: 0 м")
        self.lbl_total_length.setStyleSheet(
            "color:#f9e2af; font-family:Consolas; font-size:9pt;"
        )
        geo_l.addWidget(self.lbl_total_length)

        self.lbl_total_area = QLabel("Полигоны (0): 0 м²")
        self.lbl_total_area.setStyleSheet(
            "color:#a6e3a1; font-family:Consolas; font-size:9pt;"
        )
        geo_l.addWidget(self.lbl_total_area)

        unit_row = QWidget(); unit_rl = QHBoxLayout(unit_row)
        unit_rl.setContentsMargins(0, 0, 0, 0); unit_rl.setSpacing(6)
        self.radio_m2 = QRadioButton("м²")
        self.radio_ha = QRadioButton("га")
        self.radio_m2.setChecked(True)
        unit_btn_grp = QButtonGroup(self)
        unit_btn_grp.addButton(self.radio_m2)
        unit_btn_grp.addButton(self.radio_ha)
        self.radio_m2.toggled.connect(self._on_area_unit_changed)
        unit_rl.addWidget(QLabel("Площадь в:"))
        unit_rl.addWidget(self.radio_m2)
        unit_rl.addWidget(self.radio_ha)
        unit_rl.addStretch()
        geo_l.addWidget(unit_row)

        self.geo_g.setVisible(False)
        pl.addWidget(self.geo_g)

        # ─── Координаты курсора ───────────────────────────────────
        coord_g = QGroupBox("Координаты")
        coord_g.setStyleSheet(_gs)
        coord_l = QVBoxLayout(coord_g)
        coord_l.setContentsMargins(6, 10, 6, 6)
        coord_l.setSpacing(4)

        self.lbl_cursor = QLabel("—")
        self.lbl_cursor.setStyleSheet(
            "color:#cdd6f4; font-family:Consolas; font-size:9pt;"
        )
        coord_l.addWidget(self.lbl_cursor)

        self.lbl_last = QLabel("Последний клик: —")
        self.lbl_last.setStyleSheet(
            "color:#a6e3a1; font-family:Consolas; font-size:9pt;"
        )
        coord_l.addWidget(self.lbl_last)

        self.btn_copy = QPushButton("Скопировать в буфер")
        self.btn_copy.clicked.connect(self._copy_last_coords)
        coord_l.addWidget(self.btn_copy)
        coord_l.addStretch()

        pl.addWidget(coord_g)

        self._left_l.addWidget(panel)

        # Правая панель: список объектов
        self._main_splitter.addWidget(_left_w)
        self._main_splitter.addWidget(self._build_obj_panel())
        self._main_splitter.setSizes([10000, 2800])
        root.addWidget(self._main_splitter)

    # ── Строка UI для одного слоя ────────────────────────────────

    # ── Панель объектов (список полигонов / точек) ───────────────

    def _build_obj_panel(self) -> QWidget:
        w = QWidget()
        w.setMinimumWidth(0)
        w.setStyleSheet("QWidget { background:#181825; color:#cdd6f4; }")
        vl = QVBoxLayout(w)
        vl.setContentsMargins(4, 4, 4, 4)
        vl.setSpacing(4)

        # Заголовок
        hdr_w = QWidget()
        hdr_l = QHBoxLayout(hdr_w)
        hdr_l.setContentsMargins(0, 0, 0, 0)
        hdr_l.setSpacing(4)
        lbl_title = QLabel("Объекты карты")
        lbl_title.setStyleSheet("color:#cba6f7; font-weight:bold; font-size:9pt;")
        btn_refresh = QPushButton("↻")
        btn_refresh.setFixedSize(22, 22)
        btn_refresh.setToolTip("Обновить список")
        btn_refresh.clicked.connect(self._refresh_obj_list)
        hdr_l.addWidget(lbl_title)
        hdr_l.addStretch()
        hdr_l.addWidget(btn_refresh)
        vl.addWidget(hdr_w)

        # Дерево объектов
        self.obj_tree = QTreeWidget()
        self.obj_tree.setColumnCount(2)
        self.obj_tree.setHeaderLabels(["Элемент", "Данные"])
        self.obj_tree.setStyleSheet(
            "QTreeWidget { background:#1e1e2e; border:1px solid #313244; "
            "color:#cdd6f4; font-size:8pt; } "
            "QTreeWidget::item:selected { background:#313244; } "
            "QHeaderView::section { background:#181825; color:#a6adc8; "
            "border:none; padding:2px; font-size:8pt; }"
        )
        self.obj_tree.header().setStretchLastSection(True)
        self.obj_tree.itemClicked.connect(self._on_obj_tree_clicked)
        vl.addWidget(self.obj_tree, 1)

        return w

    def _refresh_obj_list(self):
        """Перестраивает дерево объектов по текущему состоянию холста."""
        if not hasattr(self, "obj_tree"):
            return
        tree  = self.obj_tree
        c     = self.canvas
        tree.clear()
        tree.setColumnCount(2)

        _ICON_POLY = "📐"
        _ICON_PT   = "●"
        _ICON_EDGE = "↔"

        # ─── Полигоны ─────────────────────────────────────────────
        poly_pt_ids: set = set()
        for i, poly in enumerate(c.map_polygons):
            area     = c._effective_polygon_area(poly)
            area_str = (f"{area/10000:.4f} га"
                        if c.area_unit == "ha" else f"{area:.1f} м²")
            poly_item = QTreeWidgetItem(tree,
                [f"{_ICON_POLY} Полигон {i + 1}", area_str])
            poly_item.setData(0, Qt.ItemDataRole.UserRole, ("poly", poly.id))
            poly_item.setForeground(0, QColor("#a6e3a1"))
            poly_item.setForeground(1, QColor("#f9e2af"))

            ids = poly.point_ids
            n   = len(ids)
            for j, pt_id in enumerate(ids):
                pt      = c._get_point_by_id(pt_id)
                next_pt = c._get_point_by_id(ids[(j + 1) % n])
                if pt is None:
                    continue
                poly_pt_ids.add(pt_id)
                name     = pt.label or pt_id[:6]
                coord_s  = f"{pt.lat:.6f}°N  {pt.lon:.6f}°E"
                if next_pt is not None:
                    dist  = haversine_distance(pt.lat, pt.lon,
                                               next_pt.lat, next_pt.lon)
                    bear  = c._calc_bearing(pt.lat, pt.lon,
                                             next_pt.lat, next_pt.lon)
                    d_str = (f"{dist/1000:.3f} км"
                             if dist >= 1000 else f"{dist:.1f} м")
                    meta = []
                    if c.show_angle_labels:
                        meta.append(f"→{bear:.1f}°")
                    if c.show_edge_length_labels:
                        meta.append(d_str)
                    if c.show_coord_labels:
                        info = f"{coord_s}   {'  '.join(meta)}".rstrip()
                    else:
                        info = "  ".join(meta) if meta else "—"
                else:
                    info = coord_s if c.show_coord_labels else "—"
                pt_item = QTreeWidgetItem(poly_item,
                    [f"  {_ICON_PT} {name}", info])
                pt_item.setData(0, Qt.ItemDataRole.UserRole, ("point", pt_id))
                pt_item.setForeground(0, QColor("#89b4fa"))
                pt_item.setForeground(1, QColor("#a6adc8"))
            poly_item.setExpanded(True)

        # ─── Точки без полигона ────────────────────────────────────
        lonely = [p for p in c.map_points if p.id not in poly_pt_ids]
        if lonely:
            pts_root = QTreeWidgetItem(tree,
                [f"{_ICON_PT} Точки ({len(lonely)})", ""])
            pts_root.setForeground(0, QColor("#89b4fa"))
            for pt in lonely:
                name = pt.label or pt.id[:6]
                info = f"{pt.lat:.6f}°N  {pt.lon:.6f}°E" if c.show_coord_labels else "—"
                chi  = QTreeWidgetItem(pts_root,
                    [f"  {_ICON_PT} {name}", info])
                chi.setData(0, Qt.ItemDataRole.UserRole, ("point", pt.id))
                chi.setForeground(1, QColor("#a6adc8"))
            pts_root.setExpanded(True)

        # ─── Рёбра ────────────────────────────────────────────────
        if c.map_edges:
            edge_root = QTreeWidgetItem(tree,
                [f"{_ICON_EDGE} Рёбра ({len(c.map_edges)})", ""])
            edge_root.setForeground(0, QColor("#f9e2af"))
            for e in c.map_edges:
                a = c._get_point_by_id(e.point_a_id)
                b = c._get_point_by_id(e.point_b_id)
                if a is None or b is None:
                    continue
                dist  = haversine_distance(a.lat, a.lon, b.lat, b.lon)
                bear  = c._calc_bearing(a.lat, a.lon, b.lat, b.lon)
                d_str = f"{dist/1000:.3f} км" if dist >= 1000 else f"{dist:.1f} м"
                na    = a.label or a.id[:4]
                nb    = b.label or b.id[:4]
                edge_meta = []
                if c.show_edge_length_labels:
                    edge_meta.append(d_str)
                if c.show_angle_labels:
                    edge_meta.append(f"∠{bear:.1f}°")
                chi   = QTreeWidgetItem(edge_root,
                    [f"  {na} – {nb}", "  ".join(edge_meta) if edge_meta else "—"])
                chi.setForeground(1, QColor("#a6adc8"))

        tree.resizeColumnToContents(0)

    def _on_obj_tree_clicked(self, item: QTreeWidgetItem, _col: int):
        data = item.data(0, Qt.ItemDataRole.UserRole)
        if not data:
            return
        kind, obj_id = data
        c = self.canvas
        if kind == "point":
            pt = c._get_point_by_id(obj_id)
            if pt:
                c.selected_point_ids = {obj_id}
                c.selection_changed.emit([pt])
                # Центрируем карту на точку
                c.set_center(pt.lat, pt.lon)
                c.update()
        elif kind == "poly":
            poly = next((p for p in c.map_polygons if p.id == obj_id), None)
            if poly:
                c.selected_point_ids = set(poly.point_ids)
                c.selection_changed.emit(c._get_selected_points())
                c.update()

    def _make_layer_row(self, layer: LayerInfo) -> QWidget:
        """[✓ Имя слоя]  [══════ slider ══════]  [87%]"""
        row = QWidget()
        rl  = QHBoxLayout(row)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.setSpacing(6)

        chk = QCheckBox(layer.name)
        chk.setChecked(layer.visible)
        chk.setMinimumWidth(110)
        rl.addWidget(chk)

        slider = QSlider(Qt.Orientation.Horizontal)
        slider.setRange(0, 100)
        slider.setValue(int(layer.opacity * 100))
        slider.setFixedWidth(90)
        slider.setFixedHeight(20)
        rl.addWidget(slider)

        lbl_pct = QLabel(f"{int(layer.opacity * 100)}%")
        lbl_pct.setStyleSheet("color:#a6adc8; font-size:8pt;")
        lbl_pct.setFixedWidth(30)
        rl.addWidget(lbl_pct)

        def on_chk(state, _l=layer):
            _l.visible = bool(state)
            self.canvas.update()

        def on_slider(val, _l=layer, _lbl=lbl_pct):
            _l.opacity = val / 100.0
            _lbl.setText(f"{val}%")
            self.canvas.update()

        chk.stateChanged.connect(on_chk)
        slider.valueChanged.connect(on_slider)

        return row

    # ── Загрузка баз данных ──────────────────────────────────────

    def _load_databases(self):
        osm_dir = "osm_map"
        if not os.path.isdir(osm_dir):
            return

        found = sorted(
            f for f in os.listdir(osm_dir) if f.lower().endswith(".sqlitedb")
        )

        configs: List[Tuple[str, str, float, int]] = []
        for fname in found:
            matched = False
            for pat, disp, opac, order in _LAYER_PATTERNS:
                if pat.lower() in fname.lower():
                    configs.append((fname, disp, opac, order))
                    matched = True
                    break
            if not matched:
                base_name = os.path.splitext(fname)[0]
                configs.append((fname, f"Оригинал: {base_name}", 1.0, 1))

        configs.sort(key=lambda c: c[3])   # bottom → top по полю order

        for fname, display_name, default_opacity, _ in configs:
            path = os.path.join(osm_dir, fname)
            try:
                prov  = TileProvider(path, display_name)
                layer = LayerInfo(prov, display_name, default_opacity, True)
                self.canvas.layers.append(layer)
                self._providers.append(prov)

                row_w = self._make_layer_row(layer)
                self.layers_vl.addWidget(row_w)
            except Exception as e:
                print(f"[MapTab] Ошибка загрузки {fname}: {e}")

        self.layers_vl.addStretch()

        if self.canvas.layers:
            self._center_on_coverage()
            self._sync_zoom_label()

    # ── Поиск координат ──────────────────────────────────────────

    def _on_search(self):
        text = self.search_edit.text().strip()
        if not text:
            return

        clean = text.replace("°", "").replace(",", " ").strip()
        parts = clean.split()
        try:
            if len(parts) >= 2:
                lat = float(parts[0])
                lon = float(parts[1])
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    self.canvas.set_center(lat, lon)
                    if self.chk_search_add_point.isChecked():
                        self.canvas._add_map_point_at(lat, lon)
                    self._sync_zoom_label()
                    self.search_edit.setStyleSheet("")
                    return
        except ValueError:
            pass

        self.search_edit.setStyleSheet("border: 1px solid #f38ba8;")

    # ── Переключение режима ───────────────────────────────────────

    def _on_mode_changed(self, checked: bool):
        is_view = self.radio_view.isChecked()
        is_pts  = self.radio_pts.isChecked()

        self.canvas.view_mode  = is_view
        self.canvas.point_mode = is_pts

        self.insert_g.setVisible(not is_view and not is_pts)
        self.measure_g.setVisible(is_view)
        self.points_g.setVisible(is_pts)
        self.edges_g.setVisible(is_pts)
        self.polygon_g.setVisible(is_pts)
        self.geo_g.setVisible(is_pts)

        if not is_view:
            # Очищаем временные точки при выходе из режима обзора
            self.canvas.temp_points.clear()
            self.canvas.update()
            self.lbl_distance.setText("Точек: 0")

        if not is_pts:
            # Снимаем выделение при выходе из режима точек
            self.canvas.selected_point_ids.clear()
            self.canvas._hovered_point_id  = None
            self.canvas._hovered_edge_id   = None
            self.canvas.update()
        else:
            self._update_pts_count()
            self.canvas.setFocus()   # H-key должен работать сразу

    # ── Обработчики событий ──────────────────────────────────────

    def _on_click(self, lat: float, lon: float):
        """Клик на карте в режиме редактирования — вставляем координаты в таблицу."""
        self._last_lat = lat
        self._last_lon = lon

        lat_s = f"{lat:.6f}"
        lon_s = f"{lon:.6f}"
        self.lbl_last.setText(f"Клик:  {lat_s}°N   {lon_s}°E")

        row_idx = self._find_target_row()
        if row_idx is not None:
            self._set_coords_in_table(row_idx, lat_s, lon_s)
            self._next_row_idx = row_idx + 1
            self._update_insert_status()
            self._refresh_markers()

        QApplication.clipboard().setText(f"{lat_s}, {lon_s}")

    def _on_hover(self, lat: float, lon: float):
        self.lbl_cursor.setText(f"{lat:.6f}°N   {lon:.6f}°E")
        self._sync_zoom_label()

    def _on_insert_mode_changed(self, checked: bool):
        self._insert_sequential = self.radio_seq.isChecked()
        self._update_insert_status()

    # ── Логика вставки ───────────────────────────────────────────

    def _find_target_row(self) -> Optional[int]:
        if self._insert_sequential:
            rows = self.data_tab.get_rows()
            for i in range(self._next_row_idx, len(rows)):
                if not rows[i].get("широта", "").strip():
                    return i
            for i in range(0, self._next_row_idx):
                if not rows[i].get("широта", "").strip():
                    return i
            if rows:
                return self._next_row_idx % len(rows)
            return None
        else:
            selected = self.data_tab.table.selectedIndexes()
            if selected:
                return selected[0].row()
            return None

    def _set_coords_in_table(self, row_idx: int, lat: str, lon: str):
        from core.data_manager import COLUMNS
        from PyQt6.QtWidgets import QTableWidgetItem

        table = self.data_tab.table
        if row_idx >= table.rowCount():
            return

        lat_col = COLUMNS.index("широта")
        lon_col = COLUMNS.index("долгота")

        table.blockSignals(True)
        lat_item = QTableWidgetItem(lat)
        lon_item = QTableWidgetItem(lon)
        lat_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        lon_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        table.setItem(row_idx, lat_col, lat_item)
        table.setItem(row_idx, lon_col, lon_item)
        table.blockSignals(False)

        table.scrollToItem(table.item(row_idx, lat_col))
        table.selectRow(row_idx)

    def _reset_sequential(self):
        self._next_row_idx = 0
        self._update_insert_status()

    def _update_insert_status(self):
        if self._insert_sequential:
            rows = self.data_tab.get_rows()
            for i in range(self._next_row_idx, len(rows)):
                if not rows[i].get("широта", "").strip():
                    self.lbl_insert_status.setText(f"Следующая строка: {i + 1}")
                    return
            self.lbl_insert_status.setText(f"Все строки заполнены ({len(rows)})")
        else:
            selected = self.data_tab.table.selectedIndexes()
            if selected:
                self.lbl_insert_status.setText(
                    f"Выбрана строка: {selected[0].row() + 1}"
                )
            else:
                self.lbl_insert_status.setText("Строка не выбрана")

    def _refresh_markers(self):
        rows = self.data_tab.get_rows()
        self.canvas.update_markers(rows)
        self._update_insert_status()

    def _copy_last_coords(self):
        if self._last_lat is not None and self._last_lon is not None:
            text = f"{self._last_lat:.6f}, {self._last_lon:.6f}"
            QApplication.clipboard().setText(text)
            self.lbl_last.setText(f"Скопировано: {text}")

    # ── Измерение расстояний ─────────────────────────────────────

    def _update_distance_label(self):
        """Пересчитывает и отображает суммарное расстояние по временным точкам."""
        pts = self.canvas.temp_points
        n   = len(pts)
        if n < 2:
            self.lbl_distance.setText(f"Точек: {n}")
            return
        total = sum(
            haversine_distance(pts[i][0], pts[i][1], pts[i + 1][0], pts[i + 1][1])
            for i in range(n - 1)
        )
        if total >= 1000:
            self.lbl_distance.setText(f"Итого: {total / 1000:.3f} км  ({n} т.)")
        else:
            self.lbl_distance.setText(f"Итого: {total:.1f} м  ({n} т.)")

    def _clear_temp_points(self):
        self.canvas.temp_points.clear()
        self.canvas.temp_pts_changed.emit()
        self.canvas.update()
        self.lbl_distance.setText("Точек: 0")

    # ── Навигация ────────────────────────────────────────────────

    def _center_on_coverage(self):
        if not self._providers:
            return
        center = self._providers[0].coverage_center(12)
        if center:
            self.canvas.set_center(center[0], center[1], 12)
            self._sync_zoom_label()

    def _sync_zoom_label(self):
        self.zoom_lbl.setText(str(self.canvas.zoom))

    # ── Система точек: обработчики сигналов ─────────────────────

    def _on_point_changed(self, pt: MapPoint):
        """Вызывается при добавлении или перемещении точки."""
        self._update_pts_count()
        self._fill_point_props(pt)
        self._sync_polygon_props()
        self._update_geo_labels()

    def _on_point_deleted(self, pt: MapPoint):
        self._update_pts_count()
        self._clear_point_props()
        self._update_geo_labels()

    def _on_selection_changed(self, pts: list):
        """pts — список выбранных MapPoint (может быть пустым)."""
        if not pts:
            self._clear_point_props()
        else:
            self._fill_point_props(pts[0])
        # Обновляем подпись pivot-точки
        pivot_id = self.canvas._pivot_point_id
        if pivot_id:
            piv_pt = self.canvas._get_point_by_id(pivot_id)
            self.lbl_pivot.setText(f"Опора: {piv_pt.label or pivot_id[:4]}" if piv_pt else "Опора: —")
        else:
            self.lbl_pivot.setText("Опора: —")
        self._update_geo_labels()

    # ── Система рёбер: обработчики ────────────────────────────────

    def _set_edge_color_btn(self, color_hex: str):
        self.edge_color_btn.setStyleSheet(
            f"QPushButton {{ background-color:{color_hex}; border:1px solid #45475a; "
            f"border-radius:3px; }} "
            f"QPushButton:hover {{ border:1px solid #cba6f7; }}"
        )

    def _on_edge_type_changed(self, idx: int):
        self.canvas.default_edge_type = EDGE_TYPES[idx]

    def _pick_edge_color(self):
        initial = QColor(self.canvas.default_edge_color)
        color   = QColorDialog.getColor(initial, self, "Цвет ребра")
        if color.isValid():
            self.canvas.default_edge_color = color.name()
            self._set_edge_color_btn(color.name())

    def _on_edge_width_changed(self, val: int):
        width = val / 2.0
        self.edge_width_lbl.setText(f"{width:.1f}")
        self.canvas.default_edge_width = width

    def _connect_selected_points(self):
        self.canvas.connect_selected_points()
        if self.chk_auto_name_points.isChecked():
            pts = self.canvas._get_selected_points()
            for idx, pt in enumerate(pts, start=1):
                pt.label = str(idx)
        if len(self.canvas._get_selected_points()) >= 3:
            self.canvas.create_polygon_from_selected()
        self._sync_polygon_props()
        self._update_geo_labels()

    def _disconnect_selected_points(self):
        self.canvas.delete_edges_for_selection()
        self._update_geo_labels()

    def _on_connect_mode_toggled(self, checked: bool):
        self.canvas._connect_mode = checked
        if not checked:
            self.canvas._connect_source_id = None
            self.canvas.update()

    # ── Система полигонов: обработчики ───────────────────────────

    def _set_poly_fill_btn(self, color_hex: str):
        self.poly_fill_btn.setStyleSheet(
            f"QPushButton {{ background-color:{color_hex}; border:1px solid #45475a; "
            f"border-radius:3px; }} "
            f"QPushButton:hover {{ border:1px solid #cba6f7; }}"
        )

    def _set_poly_border_btn(self, color_hex: str):
        self.poly_border_btn.setStyleSheet(
            f"QPushButton {{ background-color:{color_hex}; border:1px solid #45475a; "
            f"border-radius:3px; }} "
            f"QPushButton:hover {{ border:1px solid #cba6f7; }}"
        )

    def _sync_polygon_props(self):
        poly = self.canvas._get_selected_polygon()
        if poly is None:
            self._set_poly_fill_btn(self.canvas.default_poly_fill)
            self.poly_fill_slider.blockSignals(True)
            self.poly_fill_slider.setValue(int(round(self.canvas.default_poly_fill_opacity * 100)))
            self.poly_fill_slider.blockSignals(False)
            self.poly_fill_opacity_lbl.setText(f"{int(round(self.canvas.default_poly_fill_opacity * 100))}%")
            self._set_poly_border_btn(self.canvas.default_poly_border)
            self.poly_border_type_combo.blockSignals(True)
            self.poly_border_type_combo.setCurrentIndex(EDGE_TYPES.index(self.canvas.default_poly_border_type))
            self.poly_border_type_combo.blockSignals(False)
            return
        self._set_poly_fill_btn(poly.fill_color)
        self.poly_fill_slider.blockSignals(True)
        self.poly_fill_slider.setValue(int(round(poly.fill_opacity * 100)))
        self.poly_fill_slider.blockSignals(False)
        self.poly_fill_opacity_lbl.setText(f"{int(round(poly.fill_opacity * 100))}%")
        self._set_poly_border_btn(poly.border_color)
        self.poly_border_type_combo.blockSignals(True)
        self.poly_border_type_combo.setCurrentIndex(EDGE_TYPES.index(poly.border_type))
        self.poly_border_type_combo.blockSignals(False)

    def _pick_poly_fill_color(self):
        target = self.canvas._get_selected_polygon()
        base = target.fill_color if target else self.canvas.default_poly_fill
        color = QColorDialog.getColor(QColor(base), self, "Цвет заливки")
        if color.isValid():
            self._set_poly_fill_btn(color.name())
            if target:
                target.fill_color = color.name()
            else:
                self.canvas.default_poly_fill = color.name()
            self.canvas.update()

    def _on_poly_fill_opacity_changed(self, val: int):
        self.poly_fill_opacity_lbl.setText(f"{val}%")
        opacity = val / 100.0
        target = self.canvas._get_selected_polygon()
        if target:
            target.fill_opacity = opacity
        else:
            self.canvas.default_poly_fill_opacity = opacity
        self.canvas.update()

    def _on_poly_border_type_changed(self, idx: int):
        target = self.canvas._get_selected_polygon()
        if target:
            target.border_type = EDGE_TYPES[idx]
            self.canvas.update()
        else:
            self.canvas.default_poly_border_type = EDGE_TYPES[idx]

    def _pick_poly_border_color(self):
        target = self.canvas._get_selected_polygon()
        base = target.border_color if target else self.canvas.default_poly_border
        color = QColorDialog.getColor(QColor(base), self, "Цвет рамки")
        if color.isValid():
            self._set_poly_border_btn(color.name())
            if target:
                target.border_color = color.name()
            else:
                self.canvas.default_poly_border = color.name()
            self.canvas.update()

    def _build_labels_filter_menu(self):
        menu = QMenu(self)
        menu.setStyleSheet(
            "QMenu { background:#1e1e2e; color:#cdd6f4; border:1px solid #45475a; }"
            "QMenu::item { padding:5px 20px 5px 10px; }"
            "QMenu::item:selected { background:#313244; }"
        )
        self._label_actions = {}

        items = (
            ("global", "Показывать все метки", "show_geo_labels"),
            ("coord", "Координаты точек", "show_coord_labels"),
            ("angle", "Углы поворота", "show_angle_labels"),
            ("area", "Площади", "show_area_labels"),
            ("edge", "Длины рёбер", "show_edge_length_labels"),
        )
        for key, title, attr in items:
            act = menu.addAction(title)
            act.setCheckable(True)
            act.setChecked(bool(getattr(self.canvas, attr)))
            act.toggled.connect(lambda checked, k=key: self._on_label_filter_changed(k, checked))
            self._label_actions[key] = act

        self.btn_labels_filter.setMenu(menu)

    def _on_label_filter_changed(self, key: str, checked: bool):
        if key == "global":
            self.canvas.show_geo_labels = checked
            if checked:
                self.canvas.show_point_labels = True
                if not self.canvas.show_coord_labels and not self.canvas.show_angle_labels:
                    self.canvas.show_coord_labels = True
                if not self.canvas.show_area_labels and not self.canvas.show_edge_length_labels:
                    self.canvas.show_area_labels = True
            else:
                self.canvas.show_point_labels = False
            self._sync_label_filter_menu()
            self.canvas.update()
            self._refresh_obj_list()
            return

        mapping = {
            "coord": "show_coord_labels",
            "angle": "show_angle_labels",
            "area": "show_area_labels",
            "edge": "show_edge_length_labels",
        }
        attr = mapping.get(key)
        if not attr:
            return

        setattr(self.canvas, attr, checked)
        self.canvas.show_point_labels = self.canvas.show_coord_labels or self.canvas.show_angle_labels
        self.canvas.show_geo_labels = (
            self.canvas.show_point_labels
            or self.canvas.show_area_labels
            or self.canvas.show_edge_length_labels
        )
        self._sync_label_filter_menu()
        self.canvas.update()
        self._refresh_obj_list()

    def _sync_label_filter_menu(self):
        if not hasattr(self, "_label_actions"):
            return
        current = {
            "global": self.canvas.show_geo_labels,
            "coord": self.canvas.show_coord_labels,
            "angle": self.canvas.show_angle_labels,
            "area": self.canvas.show_area_labels,
            "edge": self.canvas.show_edge_length_labels,
        }
        for key, value in current.items():
            act = self._label_actions.get(key)
            if act is None:
                continue
            act.blockSignals(True)
            act.setChecked(bool(value))
            act.blockSignals(False)

    def _on_labels_toggled(self, checked: bool):
        """Синхронизирует меню фильтра меток при переключении через H-key."""
        _ = checked
        self._sync_label_filter_menu()
        self._refresh_obj_list()

    def _on_show_dynamic_points_toggled(self, checked: bool):
        self.canvas.show_dynamic_points = checked
        if hasattr(self, "chk_auto_dynamic_points"):
            self.chk_auto_dynamic_points.setEnabled(checked)
        self.canvas._sync_dynamic_points()
        self.canvas.update()
        self._refresh_obj_list()

    def _on_auto_dynamic_points_toggled(self, checked: bool):
        self.canvas.auto_dynamic_points = checked
        self.canvas._sync_dynamic_points()
        self.canvas.update()
        self._refresh_obj_list()

    def _create_polygon_from_selected(self):
        if self.chk_auto_name_points.isChecked():
            pts = self.canvas._get_selected_points()
            for idx, pt in enumerate(pts, start=1):
                pt.label = str(idx)
        self.canvas.create_polygon_from_selected()
        self._sync_polygon_props()
        self._update_geo_labels()

    def _split_selected_polygon(self):
        """Разрезает полигон по двум выбранным точкам."""
        sel_ids = self.canvas.selected_point_ids
        if len(sel_ids) != 2:
            return
        id_list = list(sel_ids)
        for poly in list(self.canvas.map_polygons):
            if set(poly.point_ids).issuperset(sel_ids):
                self.canvas.split_polygon(poly, id_list[0], id_list[1])
                self._update_geo_labels()
                return

    # ── Вращение ──────────────────────────────────────────────────

    def _rotate_selection(self, angle_deg: float):
        self.canvas.rotate_selection(angle_deg)
        self._update_geo_labels()

    # ── Геометрия: сводная статистика ────────────────────────────

    def _on_area_unit_changed(self):
        self.canvas.area_unit = "ha" if self.radio_ha.isChecked() else "m2"
        self.canvas.update()
        self._update_geo_labels()

    def _update_geo_labels(self):
        """Пересчитывает суммарную длину рёбер и площадь полигонов."""
        if not hasattr(self, "lbl_total_length"):
            return
        total_len = sum(
            self.canvas._calc_edge_length(e) for e in self.canvas.map_edges
        )
        if total_len >= 1000:
            len_str = f"{total_len / 1000:.3f} км"
        elif total_len >= 0.1:
            len_str = f"{total_len:.1f} м"
        else:
            len_str = "0 м"
        self.lbl_total_length.setText(f"Рёбра: {len_str}")

        polys       = self.canvas.map_polygons
        total_area  = self.canvas._calc_total_effective_area()
        if self.canvas.area_unit == "ha":
            area_str = f"{total_area / 10_000:.4f} га"
        else:
            area_str = f"{total_area:.1f} м²"
        self.lbl_total_area.setText(f"Полигоны ({len(polys)}): {area_str}")

    # ── Система точек: управление панелью свойств ────────────────

    def _fill_point_props(self, pt: MapPoint):
        """Заполняет панель свойств значениями выбранной точки."""
        self.pt_name_edit.blockSignals(True)
        self.pt_name_edit.setText(pt.label)
        self.pt_name_edit.blockSignals(False)

        icon_idx = PRESET_ICONS.index(pt.icon_type) if pt.icon_type in PRESET_ICONS else -1
        self.pt_icon_combo.blockSignals(True)
        if icon_idx >= 0:
            self.pt_icon_combo.setCurrentIndex(icon_idx)
        else:
            self.pt_icon_combo.setCurrentIndex(len(PRESET_ICONS))  # "Свой..."
        self.pt_icon_combo.blockSignals(False)

        self._set_color_btn(pt.color)

        size_val = int(round(pt.size * 10))
        self.pt_size_slider.blockSignals(True)
        self.pt_size_slider.setValue(size_val)
        self.pt_size_slider.blockSignals(False)
        self.pt_size_lbl.setText(f"{pt.size:.1f}×")

    def _clear_point_props(self):
        """Сбрасывает панель свойств к значениям по умолчанию."""
        self.pt_name_edit.blockSignals(True)
        self.pt_name_edit.clear()
        self.pt_name_edit.blockSignals(False)

        default_idx = PRESET_ICONS.index(self.canvas.default_icon_type) \
            if self.canvas.default_icon_type in PRESET_ICONS else 0
        self.pt_icon_combo.blockSignals(True)
        self.pt_icon_combo.setCurrentIndex(default_idx)
        self.pt_icon_combo.blockSignals(False)

        self._set_color_btn(self.canvas.default_color)

        size_val = int(round(self.canvas.default_size * 10))
        self.pt_size_slider.blockSignals(True)
        self.pt_size_slider.setValue(size_val)
        self.pt_size_slider.blockSignals(False)
        self.pt_size_lbl.setText(f"{self.canvas.default_size:.1f}×")

    def _update_pts_count(self):
        n = len(self.canvas.map_points)
        self.lbl_pts_count.setText(f"Точек на карте: {n}")

    def _set_color_btn(self, color_hex: str):
        """Перекрашивает кнопку выбора цвета."""
        self.pt_color_btn.setStyleSheet(
            f"QPushButton {{ background-color:{color_hex}; border:1px solid #45475a; "
            f"border-radius:3px; }} "
            f"QPushButton:hover {{ border:1px solid #cba6f7; }}"
        )

    def _get_selected_point(self) -> Optional[MapPoint]:
        """Возвращает первую выделенную точку (или None)."""
        ids = self.canvas.selected_point_ids
        if not ids:
            return None
        first_id = next(iter(ids))
        return self.canvas._get_point_by_id(first_id)

    # ── Система точек: изменение свойств ────────────────────────

    def _apply_point_name(self):
        pt = self._get_selected_point()
        name = self.pt_name_edit.text().strip()
        if pt is not None:
            pt.label = name
            poly = self.canvas._get_selected_polygon()
            if poly and name == "1":
                self.canvas._renumber_polygon_points_clockwise(poly, pt.id)
            self.canvas.update()
        else:
            pass

    def _on_icon_changed(self, idx: int):
        if idx == len(PRESET_ICONS):
            # "Свой..." — открываем файловый диалог
            self._load_custom_icon()
            return
        icon_type = PRESET_ICONS[idx]
        pt = self._get_selected_point()
        if pt is not None:
            pt.icon_type        = icon_type
            pt.custom_icon_path = ""
            self.canvas.update()
        else:
            self.canvas.default_icon_type = icon_type

    def _pick_point_color(self):
        pt      = self._get_selected_point()
        initial = QColor(pt.color if pt else self.canvas.default_color)
        color   = QColorDialog.getColor(initial, self, "Цвет точки")
        if not color.isValid():
            return
        hex_color = color.name()
        self._set_color_btn(hex_color)
        if pt is not None:
            pt.color = hex_color
            self.canvas.update()
        else:
            self.canvas.default_color = hex_color

    def _on_size_changed(self, val: int):
        size = val / 10.0
        self.pt_size_lbl.setText(f"{size:.1f}×")
        pt = self._get_selected_point()
        if pt is not None:
            pt.size = size
            self.canvas.update()
        else:
            self.canvas.default_size = size

    def _load_custom_icon(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выбрать иконку",
            "",
            "Изображения (*.png *.jpg *.jpeg *.bmp *.svg *.ico *.webp)",
        )
        if not path:
            # Отмена — возвращаем текущий тип
            pt = self._get_selected_point()
            cur_type = pt.icon_type if pt else self.canvas.default_icon_type
            idx = PRESET_ICONS.index(cur_type) if cur_type in PRESET_ICONS else 0
            self.pt_icon_combo.blockSignals(True)
            self.pt_icon_combo.setCurrentIndex(idx)
            self.pt_icon_combo.blockSignals(False)
            return

        pt = self._get_selected_point()
        if pt is not None:
            pt.icon_type        = "custom"
            pt.custom_icon_path = path
            # Инвалидируем кэш для этого пути
            self.canvas._icon_cache.pop(path, None)
            self.canvas.update()
        else:
            # Без выделенной точки — просто сохраним в дефолт (в виде иконки "circle")
            self.canvas.default_icon_type = "circle"

    # ── Импорт полигона ─────────────────────────────────────────

    @staticmethod
    def _parse_polygon_coords(raw_text: str) -> List[Tuple[float, float]]:
        """Парсит строки координат в форматах: запятая/пробел/таб/N-E/точка-разделитель."""
        points: List[Tuple[float, float]] = []
        for raw_line in raw_text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            cleaned = line.replace("°", " ")
            cleaned = re.sub(r"(?<=\d)\s*[NnСс](?=\b)", "", cleaned)
            cleaned = re.sub(r"(?<=\d)\s*[EeВв](?=\b)", "", cleaned)
            cleaned = cleaned.replace(";", " ").replace("\t", " ")

            nums = re.findall(r"[-+]?\d+(?:\.\d+)?", cleaned)
            if len(nums) < 2:
                continue

            lat = float(nums[0])
            lon = float(nums[1])
            if abs(lat) > 90 and abs(lon) <= 90:
                lat, lon = lon, lat
            points.append((lat, lon))
        return points

    @staticmethod
    def _extract_coords_from_json_node(node) -> List[Tuple[float, float]]:
        result: List[Tuple[float, float]] = []
        if isinstance(node, list):
            if len(node) >= 2 and all(isinstance(node[i], (int, float)) for i in (0, 1)):
                lat, lon = float(node[0]), float(node[1])
                if abs(lat) > 90 and abs(lon) <= 90:
                    lat, lon = lon, lat
                result.append((lat, lon))
            else:
                for item in node:
                    result.extend(MapTab._extract_coords_from_json_node(item))
        elif isinstance(node, dict):
            if "lat" in node and ("lon" in node or "lng" in node):
                lon_key = "lon" if "lon" in node else "lng"
                result.append((float(node["lat"]), float(node[lon_key])))
            elif "latitude" in node and ("longitude" in node or "lon" in node):
                lon_key = "longitude" if "longitude" in node else "lon"
                result.append((float(node["latitude"]), float(node[lon_key])))
            else:
                for value in node.values():
                    result.extend(MapTab._extract_coords_from_json_node(value))
        return result

    def _load_polygon_points_from_file(self, file_path: str) -> List[Tuple[float, float]]:
        ext = os.path.splitext(file_path)[1].lower()
        if ext in {".txt", ".csv"}:
            with open(file_path, "r", encoding="utf-8") as f:
                return self._parse_polygon_coords(f.read())

        if ext == ".json":
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return self._extract_coords_from_json_node(data)

        if ext == ".xml":
            root = ET.parse(file_path).getroot()
            points: List[Tuple[float, float]] = []
            for node in root.iter():
                lat = node.attrib.get("lat") or node.attrib.get("latitude")
                lon = (
                    node.attrib.get("lon")
                    or node.attrib.get("lng")
                    or node.attrib.get("longitude")
                )
                if lat is not None and lon is not None:
                    points.append((float(lat), float(lon)))
            if points:
                return points
            return self._parse_polygon_coords("\n".join((n.text or "") for n in root.iter()))

        if ext in {".xlsx", ".xlsm", ".xls"}:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            points: List[Tuple[float, float]] = []
            for row in ws.iter_rows(values_only=True):
                vals = [v for v in row if v is not None and str(v).strip() != ""]
                if len(vals) < 2:
                    continue
                try:
                    lat = float(str(vals[0]).replace(",", "."))
                    lon = float(str(vals[1]).replace(",", "."))
                except Exception:
                    continue
                if abs(lat) > 90 and abs(lon) <= 90:
                    lat, lon = lon, lat
                points.append((lat, lon))
            wb.close()
            return points

        return []

    def _set_polygon_import_status(self, text: str, ok: bool = False):
        if not hasattr(self, "lbl_poly_import_status"):
            return
        color = "#a6e3a1" if ok else "#f38ba8"
        self.lbl_poly_import_status.setStyleSheet(f"color:{color}; font-size:8pt;")
        self.lbl_poly_import_status.setText(text)

    def _import_polygon_from_text(self):
        text = self.poly_import_edit.toPlainText()
        points = self._parse_polygon_coords(text)
        if len(points) < 3:
            self._set_polygon_import_status("Ошибка: нужно минимум 3 валидные точки")
            return
        if self.canvas.import_polygon_points(points, auto_point_names=self.chk_auto_name_points.isChecked()):
            self._set_polygon_import_status(f"Импортировано точек: {len(points)}", ok=True)
            self._update_geo_labels()
            self._refresh_obj_list()
            self._update_pts_count()

    def _import_polygon_from_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Импорт полигона",
            "",
            "Поддерживаемые файлы (*.txt *.csv *.json *.xml *.xlsx *.xlsm *.xls);;"
            "Текст (*.txt *.csv);;JSON (*.json);;XML (*.xml);;Excel (*.xlsx *.xlsm *.xls)",
        )
        if not file_path:
            return
        try:
            points = self._load_polygon_points_from_file(file_path)
        except Exception as e:
            self._set_polygon_import_status(f"Ошибка файла: {e}")
            return
        if len(points) < 3:
            self._set_polygon_import_status("Ошибка: в файле найдено меньше 3 точек")
            return
        if self.canvas.import_polygon_points(points, auto_point_names=self.chk_auto_name_points.isChecked()):
            self.poly_import_edit.setPlainText("\n".join(f"{lat}, {lon}" for lat, lon in points))
            self._set_polygon_import_status(
                f"Импорт из файла: {os.path.basename(file_path)} ({len(points)} т.)",
                ok=True,
            )
            self._update_geo_labels()
            self._refresh_obj_list()
            self._update_pts_count()

    def _create_polygon_import_examples(self):
        folder = QFileDialog.getExistingDirectory(self, "Папка для примеров импорта")
        if not folder:
            return

        sample = [
            (55.36798, 60.21226),
            (55.36920, 60.21259),
            (55.37030, 60.21263),
            (55.37038, 60.21432),
            (55.36922, 60.21394),
            (55.36785, 60.21415),
        ]
        txt_data = "\n".join(f"{lat}, {lon}" for lat, lon in sample)
        with open(os.path.join(folder, "polygon_import_sample.txt"), "w", encoding="utf-8") as f:
            f.write(txt_data)

        with open(os.path.join(folder, "polygon_import_sample.json"), "w", encoding="utf-8") as f:
            json.dump(
                {"polygon": [{"lat": lat, "lon": lon} for lat, lon in sample]},
                f,
                ensure_ascii=False,
                indent=2,
            )

        root = ET.Element("polygon")
        for lat, lon in sample:
            ET.SubElement(root, "point", lat=f"{lat}", lon=f"{lon}")
        ET.ElementTree(root).write(
            os.path.join(folder, "polygon_import_sample.xml"),
            encoding="utf-8",
            xml_declaration=True,
        )

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Polygon"
        ws.append(["lat", "lon"])
        for lat, lon in sample:
            ws.append([lat, lon])
        wb.save(os.path.join(folder, "polygon_import_sample.xlsx"))
        wb.close()

        self._set_polygon_import_status(f"Примеры сохранены в: {folder}", ok=True)

    # ── Публичный API ────────────────────────────────────────────

    def refresh(self):
        """Вызывается при переключении на вкладку."""
        self._refresh_markers()
        self._sync_zoom_label()
